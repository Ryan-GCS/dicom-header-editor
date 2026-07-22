import streamlit as st
import pydicom
import pandas as pd
import io
import zipfile
import re
import warnings
import base64
import struct
from copy import deepcopy
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="DICOM Header Editor | SwiftMR",
    page_icon="🏥",
    layout="wide"
)

# ── Logo ─────────────────────────────────────────────
def get_image_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return None

logo_b64 = get_image_base64("logo.png")
logo_html = (
    f'<img src="data:image/png;base64,{logo_b64}" '
    f'style="width:44px;height:44px;object-fit:contain;">'
    if logo_b64 else "🏥"
)
sidebar_logo_html = (
    f'<img src="data:image/png;base64,{logo_b64}" '
    f'style="width:48px;height:48px;object-fit:contain;">'
    if logo_b64 else "🏥"
)

# ══════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════
st.markdown("""
<style>
@media (prefers-color-scheme: dark) {
    .stApp { background-color: #0f1117 !important; }
    .airs-header {
        background: linear-gradient(135deg, #1a1f2e 0%, #0d1117 100%) !important;
        border-bottom: 2px solid #00d4ff !important;
    }
    .airs-title p { color: #8892a4 !important; }
    .airs-badge {
        background: rgba(0,212,255,0.1) !important;
        border: 1px solid rgba(0,212,255,0.3) !important;
        color: #00d4ff !important;
    }
    .step-card {
        background: linear-gradient(135deg, #1a1f2e, #141820) !important;
        border: 1px solid #2a3040 !important;
    }
    .step-title { color: #e8eaf0 !important; }
    .tag-code {
        color: #00d4ff !important;
        background: rgba(0,212,255,0.08) !important;
    }
    .tag-name { color: #c8d0dc !important; }
    .tag-vr { color: #8892a4 !important; background: #1e2535 !important; }
    .tag-row { border-bottom: 1px solid #1e2535 !important; }
    .sidebar-section-title {
        color: #00d4ff !important;
        border-bottom: 1px solid #2a3040 !important;
    }
    .how-step-text { color: #c8d0dc !important; }
    .note-item { color: #c8d0dc !important; }
    .mode-label { color: #8892a4 !important; }
    .mode-label::after { background: #2a3040 !important; }
    .success-banner {
        background: linear-gradient(135deg,
            rgba(0,200,100,0.15), rgba(0,150,80,0.1)) !important;
        border: 1px solid rgba(0,200,100,0.3) !important;
        color: #00c864 !important;
    }
    .metric-card {
        background: linear-gradient(135deg, #1a1f2e, #141820) !important;
        border: 1px solid #2a3040 !important;
    }
    .metric-val { color: #e8eaf0 !important; }
    .metric-lbl { color: #8892a4 !important; }
    .mode-btn-active {
        background: linear-gradient(135deg, #00d4ff, #0066ff) !important;
        color: white !important;
    }
}
@media (prefers-color-scheme: light) {
    .stApp { background-color: #f0f4f8 !important; }
    .airs-header {
        background: linear-gradient(135deg, #ffffff 0%, #e8f0fe 100%) !important;
        border-bottom: 2px solid #0066ff !important;
    }
    .airs-title p { color: #5a6a7a !important; }
    .airs-badge {
        background: rgba(0,102,255,0.1) !important;
        border: 1px solid rgba(0,102,255,0.3) !important;
        color: #0066ff !important;
    }
    .step-card {
        background: linear-gradient(135deg, #ffffff, #f5f8ff) !important;
        border: 1px solid #d0d8e8 !important;
        box-shadow: 0 4px 24px rgba(0,0,0,0.08) !important;
    }
    .step-title { color: #1a2030 !important; }
    .tag-code {
        color: #0055cc !important;
        background: rgba(0,102,255,0.08) !important;
    }
    .tag-name { color: #2a3a4a !important; }
    .tag-vr { color: #5a6a7a !important; background: #e8eef8 !important; }
    .tag-row { border-bottom: 1px solid #d8e0f0 !important; }
    .sidebar-section-title {
        color: #0066ff !important;
        border-bottom: 1px solid #d0d8e8 !important;
    }
    .how-step-text { color: #3a4a5a !important; }
    .note-item { color: #3a4a5a !important; }
    .mode-label { color: #5a6a7a !important; }
    .mode-label::after { background: #d0d8e8 !important; }
    .success-banner {
        background: linear-gradient(135deg,
            rgba(0,180,80,0.1), rgba(0,150,60,0.08)) !important;
        border: 1px solid rgba(0,180,80,0.3) !important;
        color: #007a40 !important;
    }
    .metric-card {
        background: linear-gradient(135deg, #ffffff, #f5f8ff) !important;
        border: 1px solid #d0d8e8 !important;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important;
    }
    .metric-val { color: #1a2030 !important; }
    .metric-lbl { color: #5a6a7a !important; }
}

/* 공통 */
.airs-header {
    display: flex; align-items: center; gap: 16px;
    padding: 20px 28px;
    margin-bottom: 32px;
    border-radius: 0 0 16px 16px;
}
.airs-logo-box {
    width: 52px; height: 52px;
    background: transparent; border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
}
.airs-title h1 {
    margin: 0; font-size: 22px; font-weight: 800;
    background: linear-gradient(90deg, #00d4ff, #0066ff);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    letter-spacing: 2px;
}
.airs-title p { margin: 2px 0 0; font-size: 13px; letter-spacing: 1px; }
.airs-badge {
    margin-left: auto; padding: 6px 14px;
    border-radius: 20px; font-size: 12px;
    font-weight: 600; letter-spacing: 1px;
}
.step-card {
    border-radius: 16px; padding: 20px 24px; margin-bottom: 16px;
}
.step-header { display: flex; align-items: center; gap: 12px; }
.step-number {
    width: 36px; height: 36px;
    background: linear-gradient(135deg, #00d4ff, #0066ff);
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 16px; color: white;
    box-shadow: 0 2px 12px rgba(0,212,255,0.4); flex-shrink: 0;
}
.step-title { font-size: 18px; font-weight: 700; margin: 0; }
.success-banner {
    border-radius: 12px; padding: 14px 20px; font-weight: 600;
    display: flex; align-items: center; gap: 10px; margin: 12px 0;
}
.sidebar-section-title {
    font-size: 12px; font-weight: 700;
    letter-spacing: 1px; text-transform: uppercase;
    margin-bottom: 10px; padding-bottom: 6px;
}
.how-step { display: flex; align-items: flex-start; gap: 10px; margin-bottom: 8px; }
.how-step-num {
    width: 20px; height: 20px; flex-shrink: 0;
    background: linear-gradient(135deg, #00d4ff, #0066ff);
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 11px; font-weight: 700; color: white;
}
.how-step-text { font-size: 13px; line-height: 1.5; }
.mode-label {
    font-size: 11px; font-weight: 700;
    letter-spacing: 1px; text-transform: uppercase;
    margin: 12px 0 6px; display: flex; align-items: center; gap: 6px;
}
.mode-label::after { content: ''; flex: 1; height: 1px; }
.note-item {
    display: flex; align-items: flex-start; gap: 8px;
    margin-bottom: 8px; font-size: 13px;
}
.note-icon { font-size: 14px; flex-shrink: 0; }
.tag-row {
    display: flex; align-items: center; gap: 6px; padding: 5px 0;
}
.tag-code {
    font-family: monospace; font-size: 11px;
    padding: 2px 6px; border-radius: 4px;
    min-width: 100px; flex-shrink: 0;
}
.tag-name { font-size: 12px; flex: 1; }
.tag-vr {
    font-size: 11px; padding: 1px 5px;
    border-radius: 3px; font-family: monospace; flex-shrink: 0;
}
.metric-card {
    border-radius: 12px; padding: 16px 20px;
    text-align: center; margin-bottom: 8px;
}
.metric-val { font-size: 28px; font-weight: 800; }
.metric-lbl { font-size: 12px; font-weight: 600; letter-spacing: 1px; margin-top: 4px; }
.diff-badge {
    display: inline-block; padding: 2px 8px;
    border-radius: 10px; font-size: 11px; font-weight: 700;
}
.diff-diff  { background: rgba(255,160,0,0.18);  color: #ff9800; }
.diff-only_a{ background: rgba(33,150,243,0.15); color: #2196f3; }
.diff-only_b{ background: rgba(156,39,176,0.15); color: #9c27b0; }
.diff-match { background: rgba(76,175,80,0.12);  color: #4caf50; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════
st.markdown(f"""
<div class="airs-header">
    <div class="airs-logo-box">{logo_html}</div>
    <div class="airs-title">
        <h1>SwiftMR</h1>
        <p>DICOM Header Editor &nbsp;·&nbsp; Internal Tool</p>
    </div>
    <div class="airs-badge">v2.0</div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════
def parse_dicom(file_bytes: bytes):
    return pydicom.dcmread(io.BytesIO(file_bytes), force=True)


def is_dicom_bytes(data: bytes) -> bool:
    """DICOM magic bytes 확인 (offset 128~132 = DICM)"""
    if len(data) < 132:
        return False
    return data[128:132] == b"DICM"


def is_valid_dicom_ds(ds) -> bool:
    return len(ds) >= 3


def extract_tags(ds) -> list:
    rows = []
    for elem in ds:
        try:
            tag_str = f"({elem.tag.group:04X},{elem.tag.element:04X})"
            value = (
                str(elem.value)
                if not isinstance(elem.value, bytes)
                else f"[Binary {len(elem.value)} bytes]"
            )
            rows.append({
                "Tag": tag_str,
                "Keyword": elem.keyword if not elem.tag.is_private else "Private Tag",
                "VR": str(elem.VR),
                "Value": value,
                "Private": elem.tag.is_private,
            })
        except Exception:
            continue
    return rows


def apply_modifications_to_ds(ds, modifications: dict):
    ds_copy = deepcopy(ds)
    results = []
    for tag_str, new_value in modifications.items():
        try:
            group, element = tag_str.strip("()").split(",")
            tag = pydicom.tag.Tag(int(group, 16), int(element, 16))
            if tag in ds_copy:
                vr = ds_copy[tag].VR
                old_value = str(ds_copy[tag].value)
                if vr in ["DS", "FL", "FD"]:
                    ds_copy[tag].value = float(new_value)
                elif vr in ["IS", "SL", "SS", "UL", "US"]:
                    ds_copy[tag].value = int(new_value)
                else:
                    ds_copy[tag].value = new_value
                results.append({
                    "Tag": tag_str,
                    "Keyword": ds_copy[tag].keyword,
                    "Before": old_value,
                    "After": new_value,
                    "Status": "✅ Success",
                })
            else:
                results.append({
                    "Tag": tag_str, "Keyword": "-",
                    "Before": "-", "After": new_value,
                    "Status": "⚠️ Not found",
                })
        except Exception as e:
            results.append({
                "Tag": tag_str, "Keyword": "-",
                "Before": "-", "After": new_value,
                "Status": f"❌ Error: {e}",
            })
    output = io.BytesIO()
    ds_copy.save_as(output, write_like_original=True)
    return output.getvalue(), results


def process_zip(zip_bytes: bytes, modifications: dict):
    input_zip = zipfile.ZipFile(io.BytesIO(zip_bytes))
    output_buf = io.BytesIO()
    output_zip = zipfile.ZipFile(output_buf, "w", zipfile.ZIP_DEFLATED)
    all_results, success_count, skip_count, error_count = [], 0, 0, 0
    file_list = [f for f in input_zip.namelist() if not f.endswith("/")]
    skip_exts = {".jpg", ".jpeg", ".png", ".gif", ".bmp",
                 ".txt", ".xml", ".json", ".pdf", ".zip"}

    for filename in file_list:
        file_bytes = input_zip.read(filename)
        ext = Path(filename).suffix.lower()
        if ext in skip_exts:
            output_zip.writestr(filename, file_bytes)
            skip_count += 1
            continue
        try:
            ds = pydicom.dcmread(io.BytesIO(file_bytes), force=True)
            if len(ds) < 3:
                output_zip.writestr(filename, file_bytes)
                skip_count += 1
                continue
            modified_bytes, results = apply_modifications_to_ds(ds, modifications)
            for r in results:
                r["File"] = filename
            all_results.extend(results)
            output_zip.writestr(filename, modified_bytes)
            success_count += 1
        except Exception as e:
            error_count += 1
            all_results.append({
                "File": filename, "Tag": "-", "Keyword": "-",
                "Before": "-", "After": "-",
                "Status": f"❌ Failed: {e}",
            })
            output_zip.writestr(filename, file_bytes)

    output_zip.close()
    return output_buf.getvalue(), all_results, {
        "total": len(file_list),
        "success": success_count,
        "skipped": skip_count,
        "errors": error_count,
    }


# ── Compare Utilities ────────────────────────────────
PROTECTED_TAGS = {"(7FE0,0010)"}  # Pixel Data

def get_dcm_list_from_zip(zip_bytes: bytes) -> list[str]:
    """ZIP 내 유효한 DICOM 파일 목록 반환 (macOS 메타파일 제외)"""
    result = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if name.endswith("/"):
                continue
            base = Path(name).name
            if base.startswith("._") or base.startswith("."):
                continue
            ext = Path(name).suffix.lower()
            if ext in {".jpg", ".jpeg", ".png", ".gif", ".bmp",
                       ".txt", ".xml", ".json", ".pdf"}:
                continue
            data = zf.read(name)
            # .dcm / .dicom 확장자 or DICOM magic bytes
            if ext in {".dcm", ".dicom"} or is_dicom_bytes(data):
                try:
                    ds = pydicom.dcmread(io.BytesIO(data), force=True)
                    if is_valid_dicom_ds(ds):
                        result.append(name)
                except Exception:
                    continue
    return sorted(result)


def read_dcm_from_zip(zip_bytes: bytes, filename: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        return zf.read(filename)


def compare_dicom(ds_a, ds_b) -> pd.DataFrame:
    """두 DICOM DS를 비교해 diff DataFrame 반환"""
    def ds_to_dict(ds):
        d = {}
        for elem in ds:
            try:
                tag_str = f"({elem.tag.group:04X},{elem.tag.element:04X})"
                kw = elem.keyword if not elem.tag.is_private else "Private Tag"
                vr = str(elem.VR)
                val = (
                    str(elem.value)
                    if not isinstance(elem.value, bytes)
                    else f"[Binary {len(elem.value)} bytes]"
                )
                d[tag_str] = {"Keyword": kw, "VR": vr, "Value": val}
            except Exception:
                continue
        return d

    dict_a = ds_to_dict(ds_a)
    dict_b = ds_to_dict(ds_b)
    all_tags = sorted(set(dict_a) | set(dict_b))

    rows = []
    for tag in all_tags:
        in_a = tag in dict_a
        in_b = tag in dict_b
        if in_a and in_b:
            kw = dict_a[tag]["Keyword"]
            vr = dict_a[tag]["VR"]
            va = dict_a[tag]["Value"]
            vb = dict_b[tag]["Value"]
            status = "match" if va == vb else "diff"
        elif in_a:
            kw = dict_a[tag]["Keyword"]
            vr = dict_a[tag]["VR"]
            va = dict_a[tag]["Value"]
            vb = ""
            status = "only_a"
        else:
            kw = dict_b[tag]["Keyword"]
            vr = dict_b[tag]["VR"]
            va = ""
            vb = dict_b[tag]["Value"]
            status = "only_b"
        rows.append({
            "Tag": tag, "Keyword": kw, "VR": vr,
            "Value A": va, "Value B": vb, "Status": status,
        })
    return pd.DataFrame(rows)


def make_excel_diff(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DICOM Diff"

    fill_map = {
        "diff":   PatternFill("solid", fgColor="FFF3E0"),
        "only_a": PatternFill("solid", fgColor="E3F2FD"),
        "only_b": PatternFill("solid", fgColor="F3E5F5"),
        "match":  PatternFill("solid", fgColor="F1F8E9"),
    }
    headers = ["Tag", "Keyword", "VR", "Value A", "Value B", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CFD8DC")

    for _, row in df.iterrows():
        ws.append([row["Tag"], row["Keyword"], row["VR"],
                   row["Value A"], row["Value B"], row["Status"]])
        fill = fill_map.get(row["Status"], PatternFill())
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════
defaults = {
    # Editor mode
    "ds": None, "tags_df": None, "modifications": {},
    "filename": "", "modified_bytes": None, "mod_results": None,
    "upload_mode": None, "zip_bytes": None, "summary": None,
    "queue_msg": None,
    # Compare mode
    "cmp_a_bytes": None, "cmp_a_name": "",
    "cmp_b_bytes": None, "cmp_b_name": "",
    "cmp_a_zip_bytes": None, "cmp_b_zip_bytes": None,
    "cmp_a_zip_list": [], "cmp_b_zip_list": [],
    "cmp_a_sel": "", "cmp_b_sel": "",
    "cmp_df": None,
    "cmp_staged": {},          # {tag: new_value}
    "cmp_result_bytes": None,
    "cmp_ds_b": None,
    # App mode
    "app_mode": "editor",
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════
# MODE SELECTOR
# ══════════════════════════════════════════════════════
col_m1, col_m2, col_m3 = st.columns([2, 1, 1])
with col_m2:
    if st.button(
        "✏️  Editor Mode",
        use_container_width=True,
        type="primary" if st.session_state.app_mode == "editor" else "secondary",
    ):
        st.session_state.app_mode = "editor"
        st.rerun()
with col_m3:
    if st.button(
        "⚖️  Compare Mode",
        use_container_width=True,
        type="primary" if st.session_state.app_mode == "compare" else "secondary",
    ):
        st.session_state.app_mode = "compare"
        st.rerun()

st.markdown("<hr style='margin:8px 0 24px;opacity:0.2;'>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# PHI WARNING
# ══════════════════════════════════════════════════════
st.markdown("""
<div style="
    background: linear-gradient(135deg,rgba(255,80,80,0.12),rgba(200,0,0,0.08));
    border: 1.5px solid rgba(255,80,80,0.5);
    border-left: 4px solid #ff4444;
    border-radius: 12px; padding: 16px 20px; margin-bottom: 8px;
">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
        <span style="font-size:20px;">🔒</span>
        <span style="font-size:15px;font-weight:800;color:#ff4444;letter-spacing:1px;">
            HIPAA &amp; GDPR WARNING
        </span>
    </div>
    <div style="font-size:13px;color:#e8eaf0;line-height:1.8;">
        ⚠️ This tool runs on <b>Streamlit Cloud (external server)</b>.<br>
        ⚠️ <b>DO NOT upload files containing PHI (Protected Health Information)</b>.<br>
        ⚠️ Uploading real patient data may violate <b>HIPAA</b> and <b>GDPR</b> regulations.<br>
        ✅ Only use <b>fully anonymized or de-identified DICOM files</b>.
    </div>
</div>
""", unsafe_allow_html=True)

phi_confirmed = st.checkbox(
    "✅ I confirm that this file does NOT contain any PHI "
    "and is fully anonymized.",
    key="phi_confirm",
)
if not phi_confirmed:
    st.warning("⛔ Please confirm the above statement before uploading any files.")
    st.stop()


# ══════════════════════════════════════════════════════
# ██████████  EDITOR MODE  ██████████
# ══════════════════════════════════════════════════════
if st.session_state.app_mode == "editor":

    # ── STEP 1 : Upload ──────────────────────────────
    st.markdown("""
    <div class="step-card">
      <div class="step-header">
        <div class="step-number">1</div>
        <p class="step-title">Upload File</p>
      </div>
    </div>""", unsafe_allow_html=True)

    upload_mode = st.radio(
        "mode", ["🗂️ Single DICOM (.dcm)", "📦 Multiple DICOMs (.zip)"],
        horizontal=True, label_visibility="collapsed",
    )

    if "Single" in upload_mode:
        uploaded = st.file_uploader("Upload a DICOM file", type=["dcm", "DCM"])
        if uploaded:
            if uploaded.name != st.session_state.filename:
                fb = uploaded.read()
                st.session_state.ds = parse_dicom(fb)
                st.session_state.tags_df = extract_tags(st.session_state.ds)
                st.session_state.filename = uploaded.name
                st.session_state.upload_mode = "single"
                st.session_state.zip_bytes = None
                st.session_state.modifications = {}
                st.session_state.modified_bytes = None
                st.session_state.mod_results = None
                st.session_state.summary = None
                st.session_state.queue_msg = None
            st.success(
                f"✅ Loaded: **{st.session_state.filename}** "
                f"— {len(st.session_state.tags_df)} tags"
            )
    else:
        uploaded = st.file_uploader("Upload a ZIP file", type=["zip"])
        if uploaded:
            if uploaded.name != st.session_state.filename:
                zb = uploaded.read()
                with zipfile.ZipFile(io.BytesIO(zb)) as zf:
                    file_list = [f for f in zf.namelist() if not f.endswith("/")]
                with st.expander(f"📂 {len(file_list)} files in ZIP", expanded=False):
                    st.dataframe(pd.DataFrame({"File": file_list}), hide_index=True)

                first_dcm_bytes = None
                with zipfile.ZipFile(io.BytesIO(zb)) as zf:
                    for fname in zf.namelist():
                        if fname.endswith("/"):
                            continue
                        fb = zf.read(fname)
                        try:
                            ds_test = pydicom.dcmread(io.BytesIO(fb), force=True)
                            if len(ds_test) >= 3:
                                first_dcm_bytes = fb
                                break
                        except Exception:
                            continue

                if first_dcm_bytes:
                    st.session_state.ds = parse_dicom(first_dcm_bytes)
                    st.session_state.tags_df = extract_tags(st.session_state.ds)
                    st.session_state.zip_bytes = zb
                    st.session_state.filename = uploaded.name
                    st.session_state.upload_mode = "zip"
                    st.session_state.modifications = {}
                    st.session_state.modified_bytes = None
                    st.session_state.mod_results = None
                    st.session_state.summary = None
                    st.session_state.queue_msg = None
                else:
                    st.error("❌ No valid DICOM files found in ZIP.")

            if st.session_state.filename:
                st.success(f"✅ ZIP loaded: **{st.session_state.filename}**")

    # ── STEP 2 : Edit ────────────────────────────────
    if st.session_state.ds is not None:

        st.markdown("""
        <div class="step-card">
          <div class="step-header">
            <div class="step-number">2</div>
            <p class="step-title">View &amp; Edit Tags</p>
          </div>
        </div>""", unsafe_allow_html=True)

        col1, col2 = st.columns([3, 1])
        with col1:
            search = st.text_input(
                "🔍 Search Tags", placeholder="e.g. PatientName or 0010,0010"
            )
        with col2:
            show_private = st.checkbox("Show Private Tags", value=True)

        df = pd.DataFrame(st.session_state.tags_df)
        if not show_private:
            df = df[df["Private"] == False]
        if search:
            esc = re.escape(search)
            mask = (
                df["Keyword"].str.contains(esc, case=False, na=False, regex=True)
                | df["Tag"].str.contains(esc, case=False, na=False, regex=True)
                | df["Value"].str.contains(esc, case=False, na=False, regex=True)
            )
            df = df[mask]

        st.caption(f"Showing **{len(df)}** tag(s)")

        col_a, col_b = st.columns([2, 3])
        with col_a:
            tag_options = [
                f"{row['Tag']}  {row['Keyword']}" for _, row in df.iterrows()
            ]
            selected_option = st.selectbox(
                "Select tag to edit", tag_options, key="tag_select"
            )
            selected_tag = selected_option.split("  ")[0] if selected_option else None

        with col_b:
            if selected_tag:
                cur = df[df["Tag"] == selected_tag]["Value"].values[0]
                default_val = st.session_state.modifications.get(selected_tag, cur)
            else:
                default_val = ""
            new_value = st.text_input("New value", value=default_val)

        if st.button("📝 Queue Change", use_container_width=True, key="queue_btn"):
            val = new_value.strip()
            if not selected_tag:
                st.session_state.queue_msg = ("error", "Please select a tag.")
            elif not val:
                st.session_state.queue_msg = ("error", "Please enter a new value.")
            else:
                st.session_state.modifications[selected_tag] = val
                st.session_state.modified_bytes = None
                st.session_state.mod_results = None
                st.session_state.summary = None
                st.session_state.queue_msg = (
                    "success", f"Queued: **{selected_tag}** → `{val}`"
                )

        if st.session_state.queue_msg:
            t, m = st.session_state.queue_msg
            if t == "success":
                st.success(m)
            else:
                st.warning(m)

        if st.session_state.modifications:
            st.markdown("#### 📋 Pending Modifications")
            atdf = pd.DataFrame(st.session_state.tags_df)
            mod_data = []
            for tag, val in st.session_state.modifications.items():
                kw = atdf[atdf["Tag"] == tag]["Keyword"].values
                ori = atdf[atdf["Tag"] == tag]["Value"].values
                mod_data.append({
                    "Tag": tag,
                    "Keyword": kw[0] if len(kw) else "Unknown",
                    "Original": ori[0] if len(ori) else "-",
                    "New Value": val,
                })
            st.dataframe(
                pd.DataFrame(mod_data), use_container_width=True, hide_index=True
            )

            c1, c2 = st.columns([2, 3])
            with c1:
                if st.button("🗑️ Clear All", use_container_width=True):
                    st.session_state.modifications = {}
                    st.session_state.modified_bytes = None
                    st.session_state.mod_results = None
                    st.session_state.summary = None
                    st.session_state.queue_msg = None
                    st.rerun()
            with c2:
                tag_to_remove = st.selectbox(
                    "Remove specific tag",
                    list(st.session_state.modifications.keys()),
                    key="remove_select",
                )
                if st.button("❌ Remove Selected", use_container_width=True):
                    del st.session_state.modifications[tag_to_remove]
                    st.session_state.modified_bytes = None
                    st.session_state.mod_results = None
                    st.session_state.queue_msg = None
                    st.rerun()

        with st.expander("📄 View All Tags", expanded=False):
            def highlight_modified(row):
                if row["Tag"] in st.session_state.modifications:
                    return ["background-color: rgba(0,212,255,0.1)"] * len(row)
                return [""] * len(row)

            st.dataframe(
                df.drop(columns=["Private"]).style.apply(highlight_modified, axis=1),
                use_container_width=True,
                height=400,
            )

        # ── STEP 3 : Download ────────────────────────
        st.markdown("""
        <div class="step-card">
          <div class="step-header">
            <div class="step-number">3</div>
            <p class="step-title">Apply &amp; Download</p>
          </div>
        </div>""", unsafe_allow_html=True)

        if not st.session_state.modifications:
            st.info("💡 No modifications queued yet.")
        else:
            st.write(f"**{len(st.session_state.modifications)} tag(s)** queued.")

            if st.session_state.upload_mode == "single":
                if st.button(
                    "🚀 Apply Changes", type="primary", use_container_width=True
                ):
                    with st.spinner("Processing..."):
                        mb, results = apply_modifications_to_ds(
                            st.session_state.ds, st.session_state.modifications
                        )
                    st.session_state.modified_bytes = mb
                    st.session_state.mod_results = results

                if st.session_state.mod_results:
                    st.dataframe(
                        pd.DataFrame(st.session_state.mod_results),
                        use_container_width=True, hide_index=True,
                    )
                if st.session_state.modified_bytes:
                    st.markdown(
                        '<div class="success-banner">✅ File ready for download!</div>',
                        unsafe_allow_html=True,
                    )
                    st.download_button(
                        label="⬇️ Download Modified DICOM",
                        data=st.session_state.modified_bytes,
                        file_name=f"modified_{st.session_state.filename}",
                        mime="application/octet-stream",
                        use_container_width=True,
                    )

            elif st.session_state.upload_mode == "zip":
                st.info("📦 All DICOM files in ZIP will be modified.")
                if st.button(
                    "🚀 Apply to All & Create ZIP",
                    type="primary", use_container_width=True,
                ):
                    with st.spinner("Processing ZIP... Please wait."):
                        rb, all_results, summary = process_zip(
                            st.session_state.zip_bytes,
                            st.session_state.modifications,
                        )
                    st.session_state.modified_bytes = rb
                    st.session_state.mod_results = all_results
                    st.session_state.summary = summary

                if st.session_state.summary is not None:
                    s = st.session_state.summary
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("📁 Total", s["total"])
                    c2.metric("✅ Processed", s["success"])
                    c3.metric("⏭️ Skipped", s["skipped"])
                    c4.metric("❌ Errors", s["errors"])
                    if s["success"] == 0:
                        st.error("❌ No DICOM files were processed!")

                if st.session_state.mod_results:
                    with st.expander("📊 Modification Report", expanded=False):
                        st.dataframe(
                            pd.DataFrame(st.session_state.mod_results),
                            use_container_width=True, height=400, hide_index=True,
                        )

                if st.session_state.modified_bytes is not None:
                    zip_name = st.session_state.filename.replace(
                        ".zip", "_modified.zip"
                    )
                    st.markdown(
                        '<div class="success-banner">✅ ZIP ready for download!</div>',
                        unsafe_allow_html=True,
                    )
                    st.download_button(
                        label="⬇️ Download Modified ZIP",
                        data=st.session_state.modified_bytes,
                        file_name=zip_name,
                        mime="application/zip",
                        use_container_width=True,
                        type="primary",
                    )


# ══════════════════════════════════════════════════════
# ██████████  COMPARE MODE  ██████████
# ══════════════════════════════════════════════════════
else:
    st.markdown("""
    <div class="step-card">
      <div class="step-header">
        <div class="step-number">⚖️</div>
        <p class="step-title">Compare Mode — DICOM Diff Engine</p>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── STEP C-1 : Upload A & B ──────────────────────
    st.markdown("### 📁 Step 1 · Upload Files")
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("**🔵 File A — Reference**")
        up_a = st.file_uploader(
            "File A", type=["dcm", "DCM", "zip"],
            key="cmp_up_a", label_visibility="collapsed",
        )
        if up_a:
            raw = up_a.read()
            if up_a.name.lower().endswith(".zip"):
                st.session_state.cmp_a_zip_bytes = raw
                st.session_state.cmp_a_zip_list = get_dcm_list_from_zip(raw)
                st.session_state.cmp_a_bytes = None
                st.session_state.cmp_a_name = up_a.name
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}
            else:
                st.session_state.cmp_a_bytes = raw
                st.session_state.cmp_a_name = up_a.name
                st.session_state.cmp_a_zip_bytes = None
                st.session_state.cmp_a_zip_list = []
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}

        # ZIP → 파일 선택
        if st.session_state.cmp_a_zip_list:
            sel_a = st.selectbox(
                f"📂 Select from ZIP A ({len(st.session_state.cmp_a_zip_list)} files)",
                st.session_state.cmp_a_zip_list, key="cmp_a_sel_box",
            )
            if sel_a != st.session_state.cmp_a_sel:
                st.session_state.cmp_a_sel = sel_a
                st.session_state.cmp_a_bytes = read_dcm_from_zip(
                    st.session_state.cmp_a_zip_bytes, sel_a
                )
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}

        if st.session_state.cmp_a_bytes:
            st.success(f"✅ A: **{st.session_state.cmp_a_name}**"
                       + (f" › {st.session_state.cmp_a_sel}"
                          if st.session_state.cmp_a_sel else ""))

    with col_b:
        st.markdown("**🟣 File B — Target**")
        up_b = st.file_uploader(
            "File B", type=["dcm", "DCM", "zip"],
            key="cmp_up_b", label_visibility="collapsed",
        )
        if up_b:
            raw = up_b.read()
            if up_b.name.lower().endswith(".zip"):
                st.session_state.cmp_b_zip_bytes = raw
                st.session_state.cmp_b_zip_list = get_dcm_list_from_zip(raw)
                st.session_state.cmp_b_bytes = None
                st.session_state.cmp_b_name = up_b.name
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}
            else:
                st.session_state.cmp_b_bytes = raw
                st.session_state.cmp_b_name = up_b.name
                st.session_state.cmp_b_zip_bytes = None
                st.session_state.cmp_b_zip_list = []
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}

        if st.session_state.cmp_b_zip_list:
            sel_b = st.selectbox(
                f"📂 Select from ZIP B ({len(st.session_state.cmp_b_zip_list)} files)",
                st.session_state.cmp_b_zip_list, key="cmp_b_sel_box",
            )
            if sel_b != st.session_state.cmp_b_sel:
                st.session_state.cmp_b_sel = sel_b
                st.session_state.cmp_b_bytes = read_dcm_from_zip(
                    st.session_state.cmp_b_zip_bytes, sel_b
                )
                st.session_state.cmp_df = None
                st.session_state.cmp_staged = {}

        if st.session_state.cmp_b_bytes:
            st.success(f"✅ B: **{st.session_state.cmp_b_name}**"
                       + (f" › {st.session_state.cmp_b_sel}"
                          if st.session_state.cmp_b_sel else ""))

    # ── Run Compare ──────────────────────────────────
    can_compare = (
        st.session_state.cmp_a_bytes is not None
        and st.session_state.cmp_b_bytes is not None
    )

    if can_compare:
        if st.button(
            "🔍 Run Comparison", type="primary", use_container_width=True
        ):
            with st.spinner("Comparing DICOM headers..."):
                ds_a = parse_dicom(st.session_state.cmp_a_bytes)
                ds_b = parse_dicom(st.session_state.cmp_b_bytes)
                st.session_state.cmp_df = compare_dicom(ds_a, ds_b)
                st.session_state.cmp_ds_b = ds_b
                st.session_state.cmp_staged = {}
                st.session_state.cmp_result_bytes = None

    # ── STEP C-2 : Results ───────────────────────────
    if st.session_state.cmp_df is not None:
        df_full = st.session_state.cmp_df
        st.markdown("---")
        st.markdown("### 📊 Step 2 · Comparison Results")

        # 요약 메트릭
        total   = len(df_full)
        n_match = (df_full["Status"] == "match").sum()
        n_diff  = (df_full["Status"] == "diff").sum()
        n_only_a= (df_full["Status"] == "only_a").sum()
        n_only_b= (df_full["Status"] == "only_b").sum()

        mc = st.columns(5)
        metric_data = [
            ("📊 Total",     total,    "#607d8b"),
            ("✅ Identical", n_match,  "#4caf50"),
            ("⚠️ Different", n_diff,   "#ff9800"),
            ("🔵 Only A",   n_only_a, "#2196f3"),
            ("🟣 Only B",   n_only_b, "#9c27b0"),
        ]
        for col, (lbl, val, color) in zip(mc, metric_data):
            col.markdown(
                f"""<div class="metric-card">
                    <div class="metric-val" style="color:{color};">{val}</div>
                    <div class="metric-lbl">{lbl}</div>
                </div>""",
                unsafe_allow_html=True,
            )

        # 필터 & 검색
        st.markdown("#### 🔎 Filter & Search")
        f1, f2, f3 = st.columns([3, 2, 1])
        with f1:
            cmp_search = st.text_input(
                "Search", placeholder="Tag / Keyword / Value",
                key="cmp_search", label_visibility="collapsed",
            )
        with f2:
            status_filter = st.multiselect(
                "Status filter",
                ["diff", "only_a", "only_b", "match"],
                default=["diff", "only_a", "only_b"],
                key="cmp_status_filter",
                label_visibility="collapsed",
            )
        with f3:
            page_size = st.selectbox(
                "Rows", [25, 50, 100, 200], key="cmp_page_size",
                label_visibility="collapsed",
            )

        df_view = df_full[df_full["Status"].isin(status_filter)].copy()
        if cmp_search:
            esc = re.escape(cmp_search)
            mask = (
                df_view["Tag"].str.contains(esc, case=False, na=False)
                | df_view["Keyword"].str.contains(esc, case=False, na=False)
                | df_view["Value A"].str.contains(esc, case=False, na=False)
                | df_view["Value B"].str.contains(esc, case=False, na=False)
            )
            df_view = df_view[mask]

        st.caption(f"Showing **{len(df_view)}** / {total} tag(s)")

        # 색상 하이라이트
        STATUS_COLORS = {
            "diff":   "rgba(255,152,0,0.15)",
            "only_a": "rgba(33,150,243,0.12)",
            "only_b": "rgba(156,39,176,0.12)",
            "match":  "rgba(76,175,80,0.08)",
        }

        def color_rows(row):
            c = STATUS_COLORS.get(row["Status"], "")
            return [f"background-color:{c}"] * len(row)

        # 페이지네이션
        total_pages = max(1, (len(df_view) - 1) // page_size + 1)
        page = st.number_input(
            f"Page (1–{total_pages})", min_value=1,
            max_value=total_pages, value=1, key="cmp_page",
        )
        df_page = df_view.iloc[(page - 1) * page_size: page * page_size]

        st.dataframe(
            df_page.style.apply(color_rows, axis=1),
            use_container_width=True,
            height=420,
            hide_index=True,
        )

        # ── STEP C-3 : Inline Edit ───────────────────
        st.markdown("---")
        st.markdown("### ✏️ Step 3 · Inline Edit (File B)")

        editable_df = df_full[
            (df_full["Status"].isin(["diff", "only_a", "only_b"]))
            & (~df_full["Tag"].isin(PROTECTED_TAGS))
        ].copy()

        if editable_df.empty:
            st.info("🎉 No differences found — files are identical!")
        else:
            edit_options = [
                f"{row['Tag']}  {row['Keyword']}  [{row['Status']}]"
                for _, row in editable_df.iterrows()
            ]
            sel_edit = st.selectbox(
                "Select tag to edit in File B",
                edit_options, key="cmp_edit_sel",
            )
            sel_tag = sel_edit.split("  ")[0] if sel_edit else None

            if sel_tag:
                row_data = editable_df[editable_df["Tag"] == sel_tag].iloc[0]
                ec1, ec2, ec3 = st.columns([2, 2, 1])
                with ec1:
                    st.markdown(
                        f"**Value A (Ref):** `{row_data['Value A'] or '—'}`"
                    )
                with ec2:
                    current_b = st.session_state.cmp_staged.get(
                        sel_tag, row_data["Value B"]
                    )
                    new_b_val = st.text_input(
                        "New value for B", value=current_b, key="cmp_new_val"
                    )
                with ec3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("📌 Stage", use_container_width=True):
                        v = new_b_val.strip()
                        if v:
                            st.session_state.cmp_staged[sel_tag] = v
                            st.session_state.cmp_result_bytes = None
                            st.success(f"Staged: **{sel_tag}** → `{v}`")
                        else:
                            st.warning("Please enter a value.")

            # Staged 목록
            if st.session_state.cmp_staged:
                st.markdown("#### 📋 Staged Edits")
                staged_rows = []
                for tag, val in st.session_state.cmp_staged.items():
                    orig = editable_df[editable_df["Tag"] == tag]["Value B"].values
                    staged_rows.append({
                        "Tag": tag,
                        "Original B": orig[0] if len(orig) else "-",
                        "New Value": val,
                    })
                st.dataframe(
                    pd.DataFrame(staged_rows),
                    use_container_width=True, hide_index=True,
                )

                sc1, sc2 = st.columns([1, 2])
                with sc1:
                    if st.button(
                        "🗑️ Clear Staged", use_container_width=True
                    ):
                        st.session_state.cmp_staged = {}
                        st.session_state.cmp_result_bytes = None
                        st.rerun()
                with sc2:
                    if st.button(
                        "🚀 Apply to File B", type="primary",
                        use_container_width=True,
                    ):
                        with st.spinner("Applying changes..."):
                            rb, _ = apply_modifications_to_ds(
                                st.session_state.cmp_ds_b,
                                st.session_state.cmp_staged,
                            )
                        st.session_state.cmp_result_bytes = rb
                        st.success("✅ Changes applied!")

                if st.session_state.cmp_result_bytes:
                    b_base = Path(st.session_state.cmp_b_name).stem
                    st.markdown(
                        '<div class="success-banner">'
                        "✅ Modified File B ready for download!"
                        "</div>",
                        unsafe_allow_html=True,
                    )
                    st.download_button(
                        label="⬇️ Download Modified File B",
                        data=st.session_state.cmp_result_bytes,
                        file_name=f"{b_base}_modified.dcm",
                        mime="application/octet-stream",
                        use_container_width=True,
                    )

        # ── STEP C-4 : Export ────────────────────────
        st.markdown("---")
        st.markdown("### 📥 Step 4 · Export Diff Report")

        ex1, ex2 = st.columns(2)
        with ex1:
            csv_bytes = df_full.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="⬇️ Full Diff CSV",
                data=csv_bytes,
                file_name="dicom_diff.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with ex2:
            if OPENPYXL_OK:
                xlsx_bytes = make_excel_diff(df_full)
                st.download_button(
                    label="⬇️ Full Diff Excel (색상 코딩)",
                    data=xlsx_bytes,
                    file_name="dicom_diff.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )
            else:
                st.info("💡 `pip install openpyxl` to enable Excel export.")

        # Change Log CSV (staged only)
        if st.session_state.cmp_staged:
            log_rows = []
            for tag, val in st.session_state.cmp_staged.items():
                orig = df_full[df_full["Tag"] == tag]["Value B"].values
                log_rows.append({
                    "Tag": tag,
                    "Original B": orig[0] if len(orig) else "-",
                    "New Value": val,
                })
            log_csv = pd.DataFrame(log_rows).to_csv(index=False).encode("utf-8")
            st.download_button(
                label="⬇️ Change Log CSV",
                data=log_csv,
                file_name="change_log.csv",
                mime="text/csv",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:16px 0 20px;">
        <div style="width:56px;height:56px;margin:0 auto 10px;
            display:flex;align-items:center;justify-content:center;">
            {sidebar_logo_html}
        </div>
        <div style="font-size:14px;font-weight:800;letter-spacing:2px;
            background:linear-gradient(90deg,#00d4ff,#0066ff);
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;">
            SwiftMR</div>
        <div style="font-size:11px;color:#8892a4;margin-top:2px;letter-spacing:1px;">
            DICOM Header Editor</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<div class="sidebar-section-title">📖 How to Use</div>',
        unsafe_allow_html=True,
    )
    st.markdown("""
    <div class="mode-label">Editor Mode</div>
    <div class="how-step"><div class="how-step-num">1</div>
        <div class="how-step-text">Upload <code>.dcm</code> or <code>.zip</code></div></div>
    <div class="how-step"><div class="how-step-num">2</div>
        <div class="how-step-text">Select tag → Enter new value → Queue</div></div>
    <div class="how-step"><div class="how-step-num">3</div>
        <div class="how-step-text">Apply &amp; Download</div></div>
    <div class="mode-label" style="margin-top:14px;">Compare Mode</div>
    <div class="how-step"><div class="how-step-num">1</div>
        <div class="how-step-text">Upload File A (Reference) + File B (Target)</div></div>
    <div class="how-step"><div class="how-step-num">2</div>
        <div class="how-step-text">Run Comparison → View Diff</div></div>
    <div class="how-step"><div class="how-step-num">3</div>
        <div class="how-step-text">Stage edits → Apply → Download</div></div>
    <div class="how-step"><div class="how-step-num">4</div>
        <div class="how-step-text">Export CSV / Excel report</div></div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<div class="sidebar-section-title">⚠️ Notes</div>',
        unsafe_allow_html=True,
    )
    st.markdown("""
    <div class="note-item"><span class="note-icon">🔒</span>
        <span>Original files are <b>never modified</b></span></div>
    <div class="note-item"><span class="note-icon">🛡️</span>
        <span>Pixel Data <b>(7FE0,0010)</b> is protected</span></div>
    <div class="note-item"><span class="note-icon">📦</span>
        <span>ZIP: macOS meta-files auto-excluded</span></div>
    <div class="note-item"><span class="note-icon">🚫</span>
        <span>Do <b>not</b> upload real patient data (PHI)</span></div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<div class="sidebar-section-title">🔧 Common Tags</div>',
        unsafe_allow_html=True,
    )

    tab1, tab2, tab3 = st.tabs(["Standard", "AIRS Upload", "AIRS Recon"])

    with tab1:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(0020,000E)</span>
            <span class="tag-name">SeriesInstanceUID</span></div>
        <div class="tag-row"><span class="tag-code">(0020,000D)</span>
            <span class="tag-name">StudyInstanceUID</span></div>
        <div class="tag-row"><span class="tag-code">(0008,103E)</span>
            <span class="tag-name">SeriesDescription</span></div>
        <div class="tag-row"><span class="tag-code">(0010,0010)</span>
            <span class="tag-name">PatientName</span></div>
        <div class="tag-row"><span class="tag-code">(0008,0060)</span>
            <span class="tag-name">Modality</span></div>
        <div class="tag-row"><span class="tag-code">(2001,9000)</span>
            <span class="tag-name">Private (Philips)</span></div>
        """, unsafe_allow_html=True)

    with tab2:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(00E1,1010)</span>
            <span class="tag-name">StudyId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1011)</span>
            <span class="tag-name">SeriesId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1012)</span>
            <span class="tag-name">ImageId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1014)</span>
            <span class="tag-name">UploadId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1015)</span>
            <span class="tag-name">DeviceId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1030)</span>
            <span class="tag-name">DispatchUnitId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1031)</span>
            <span class="tag-name">PostProcessingType</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1032)</span>
            <span class="tag-name">SourceDispatchUnitId</span><span class="tag-vr">SH</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1033)</span>
            <span class="tag-name">ADCType</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1034)</span>
            <span class="tag-name">ADCNoiseThreshold</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1035)</span>
            <span class="tag-name">BValue</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1036)</span>
            <span class="tag-name">IsMarked</span><span class="tag-vr">CS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1037)</span>
            <span class="tag-name">NumberOfProjections</span><span class="tag-vr">IS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1038)</span>
            <span class="tag-name">RadialAngle</span><span class="tag-vr">IS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1039)</span>
            <span class="tag-name">Zoom</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1040)</span>
            <span class="tag-name">SliceOrder</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1041)</span>
            <span class="tag-name">Orientation</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1042)</span>
            <span class="tag-name">Gap</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1043)</span>
            <span class="tag-name">Thickness</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1044)</span>
            <span class="tag-name">CalculatedBvalue</span><span class="tag-vr">SL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1045)</span>
            <span class="tag-name">PostprocessingMode</span><span class="tag-vr">LO</span></div>
        """, unsafe_allow_html=True)

    with tab3:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(00E1,1020)</span>
            <span class="tag-name">InputSnr</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1021)</span>
            <span class="tag-name">OutputSnr</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1022)</span>
            <span class="tag-name">DenoisingLevel</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1023)</span>
            <span class="tag-name">Sharpness</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1024)</span>
            <span class="tag-name">ModelPath</span><span class="tag-vr">OB</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1025)</span>
            <span class="tag-name">ModelHeader</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1026)</span>
            <span class="tag-name">ModelVersion</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1034)</span>
            <span class="tag-name">ADCNoiseThreshold</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1035)</span>
            <span class="tag-name">BValue</span><span class="tag-vr">LO</span></div>
        """, unsafe_allow_html=True)

    st.divider()
    st.markdown("""
    <div style="text-align:center;font-size:11px;color:#4a5568;padding:8px 0;">
        © 2026 AIRS Medical Inc.<br>All rights reserved.<br>
        Global Technical Support
    </div>
    """, unsafe_allow_html=True)
