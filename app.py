import streamlit as st
import pydicom
import pandas as pd
import io
import zipfile
import re
import warnings
import base64
import html as html_mod
from copy import deepcopy
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="DICOM Header Editor | SwiftMR",
    page_icon="🏥",
    layout="wide"
)

# ══════════════════════════════════════════════════════
# LOGO
# ══════════════════════════════════════════════════════
def get_image_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return None

logo_b64 = get_image_base64("logo.png")
logo_html = (
    f'<img src="data:image/png;base64,{logo_b64}" '
    f'style="width:44px;height:44px;object-fit:contain;">'
    if logo_b64 else "🏥"
)
sidebar_logo_html = (
    f'<img src="data:image/png;base64,{logo_b64}" '
    f'style="width:48px;height:48px;object-fit:contain;">'
    if logo_b64 else "🏥"
)

# ══════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════
st.markdown("""
<style>
@media (prefers-color-scheme: dark) {
    .stApp { background-color: #0f1117 !important; }
    .airs-header {
        background: linear-gradient(135deg, #1a1f2e 0%, #0d1117 100%) !important;
        border-bottom: 2px solid #00d4ff !important;
    }
    .airs-title p { color: #8892a4 !important; }
    .airs-badge {
        background: rgba(0,212,255,0.1) !important;
        border: 1px solid rgba(0,212,255,0.3) !important;
        color: #00d4ff !important;
    }
    .step-card {
        background: linear-gradient(135deg, #1a1f2e, #141820) !important;
        border: 1px solid #2a3040 !important;
    }
    .step-title { color: #e8eaf0 !important; }
    .metric-card {
        background: linear-gradient(135deg, #1a1f2e, #141820) !important;
        border: 1px solid #2a3040 !important;
    }
    .metric-val { color: #e8eaf0 !important; }
    .metric-lbl { color: #8892a4 !important; }
    .diff-banner-warn {
        background: rgba(255,152,0,0.12) !important;
        border: 1.5px solid rgba(255,152,0,0.4) !important;
        border-left: 4px solid #ff9800 !important;
    }
    .diff-banner-ok {
        background: rgba(76,175,80,0.1) !important;
        border: 1.5px solid rgba(76,175,80,0.3) !important;
        border-left: 4px solid #4caf50 !important;
    }
    .file-header-a {
        background: linear-gradient(135deg,rgba(0,212,255,0.12),rgba(0,100,200,0.08)) !important;
        border: 1px solid rgba(0,212,255,0.3) !important;
    }
    .file-header-b {
        background: linear-gradient(135deg,rgba(156,39,176,0.12),rgba(100,0,180,0.08)) !important;
        border: 1px solid rgba(156,39,176,0.3) !important;
    }
    .apply-hint {
        background: rgba(0,150,255,0.1) !important;
        border: 1px solid rgba(0,150,255,0.25) !important;
        color: #60aaff !important;
    }
    .tag-code { color: #00d4ff !important; background: rgba(0,212,255,0.08) !important; }
    .tag-name { color: #c8d0dc !important; }
    .tag-vr   { color: #8892a4 !important; background: #1e2535 !important; }
    .tag-row  { border-bottom: 1px solid #1e2535 !important; }
    .sidebar-section-title {
        color: #00d4ff !important;
        border-bottom: 1px solid #2a3040 !important;
    }
    .how-step-text { color: #c8d0dc !important; }
    .note-item     { color: #c8d0dc !important; }
    .mode-label    { color: #8892a4 !important; }
    .mode-label::after { background: #2a3040 !important; }
    .success-banner {
        background: linear-gradient(135deg,rgba(0,200,100,0.15),rgba(0,150,80,0.1)) !important;
        border: 1px solid rgba(0,200,100,0.3) !important;
        color: #00c864 !important;
    }
    .dl-target-box {
        background: linear-gradient(135deg,rgba(255,152,0,0.15),rgba(200,100,0,0.1)) !important;
        border: 2px solid rgba(255,152,0,0.6) !important;
        border-left: 5px solid #ff9800 !important;
    }
    .dl-target-title { color: #ffb74d !important; }
    .dl-target-desc  { color: #c8d0dc !important; }
    .table-info-text { color: #8892a4 !important; }
    .page-ctrl-box {
        background: rgba(0,212,255,0.05) !important;
        border: 1px solid rgba(0,212,255,0.15) !important;
    }
    .page-num   { color: #00d4ff !important; }
    .page-total { color: #8892a4 !important; }
    .manual-edit-box {
        background: linear-gradient(135deg,rgba(0,200,150,0.1),rgba(0,150,100,0.06)) !important;
        border: 1.5px solid rgba(0,200,150,0.35) !important;
        border-left: 4px solid #00c896 !important;
    }
    .val-card {
        background: #1a1f2e !important;
        border: 1px solid #2a3040 !important;
    }
    .val-label { color: #8892a4 !important; }
    .val-content {
        color: #e8eaf0 !important;
        background: #0f1117 !important;
        border: 1px solid #2a3040 !important;
    }
    .val-status-match  { color: #4caf50 !important; }
    .val-status-diff   { color: #ff9800 !important; }
    .val-status-only_a { color: #2196f3 !important; }
    .val-status-only_b { color: #9c27b0 !important; }
}
@media (prefers-color-scheme: light) {
    .stApp { background-color: #f0f4f8 !important; }
    .airs-header {
        background: linear-gradient(135deg,#ffffff,#e8f0fe) !important;
        border-bottom: 2px solid #0066ff !important;
    }
    .airs-title p { color: #5a6a7a !important; }
    .airs-badge {
        background: rgba(0,102,255,0.1) !important;
        border: 1px solid rgba(0,102,255,0.3) !important;
        color: #0066ff !important;
    }
    .step-card {
        background: linear-gradient(135deg,#ffffff,#f5f8ff) !important;
        border: 1px solid #d0d8e8 !important;
        box-shadow: 0 4px 24px rgba(0,0,0,0.08) !important;
    }
    .step-title { color: #1a2030 !important; }
    .metric-card {
        background: linear-gradient(135deg,#ffffff,#f5f8ff) !important;
        border: 1px solid #d0d8e8 !important;
    }
    .metric-val { color: #1a2030 !important; }
    .metric-lbl { color: #5a6a7a !important; }
    .diff-banner-warn {
        background: rgba(255,152,0,0.08) !important;
        border: 1.5px solid rgba(255,152,0,0.35) !important;
        border-left: 4px solid #ff9800 !important;
    }
    .diff-banner-ok {
        background: rgba(76,175,80,0.08) !important;
        border: 1.5px solid rgba(76,175,80,0.3) !important;
        border-left: 4px solid #4caf50 !important;
    }
    .file-header-a {
        background: linear-gradient(135deg,rgba(0,150,255,0.08),rgba(0,80,200,0.05)) !important;
        border: 1px solid rgba(0,150,255,0.25) !important;
    }
    .file-header-b {
        background: linear-gradient(135deg,rgba(156,39,176,0.08),rgba(100,0,180,0.05)) !important;
        border: 1px solid rgba(156,39,176,0.25) !important;
    }
    .apply-hint {
        background: rgba(0,100,255,0.06) !important;
        border: 1px solid rgba(0,100,255,0.2) !important;
        color: #0055cc !important;
    }
    .tag-code { color: #0055cc !important; background: rgba(0,102,255,0.08) !important; }
    .tag-name { color: #2a3a4a !important; }
    .tag-vr   { color: #5a6a7a !important; background: #e8eef8 !important; }
    .tag-row  { border-bottom: 1px solid #d8e0f0 !important; }
    .sidebar-section-title {
        color: #0066ff !important;
        border-bottom: 1px solid #d0d8e8 !important;
    }
    .how-step-text { color: #3a4a5a !important; }
    .note-item     { color: #3a4a5a !important; }
    .mode-label    { color: #5a6a7a !important; }
    .mode-label::after { background: #d0d8e8 !important; }
    .success-banner {
        background: linear-gradient(135deg,rgba(0,180,80,0.1),rgba(0,150,60,0.08)) !important;
        border: 1px solid rgba(0,180,80,0.3) !important;
        color: #007a40 !important;
    }
    .dl-target-box {
        background: linear-gradient(135deg,rgba(255,152,0,0.1),rgba(200,100,0,0.06)) !important;
        border: 2px solid rgba(255,152,0,0.5) !important;
        border-left: 5px solid #ff9800 !important;
    }
    .dl-target-title { color: #e65100 !important; }
    .dl-target-desc  { color: #3a4a5a !important; }
    .table-info-text { color: #5a6a7a !important; }
    .page-ctrl-box {
        background: rgba(0,102,255,0.04) !important;
        border: 1px solid rgba(0,102,255,0.15) !important;
    }
    .page-num   { color: #0066ff !important; }
    .page-total { color: #5a6a7a !important; }
    .manual-edit-box {
        background: linear-gradient(135deg,rgba(0,180,130,0.08),rgba(0,150,100,0.05)) !important;
        border: 1.5px solid rgba(0,180,130,0.3) !important;
        border-left: 4px solid #00a878 !important;
    }
    .val-card {
        background: #ffffff !important;
        border: 1px solid #d0d8e8 !important;
    }
    .val-label { color: #5a6a7a !important; }
    .val-content {
        color: #1a2030 !important;
        background: #f5f8ff !important;
        border: 1px solid #d0d8e8 !important;
    }
    .val-status-match  { color: #2e7d32 !important; }
    .val-status-diff   { color: #e65100 !important; }
    .val-status-only_a { color: #1565c0 !important; }
    .val-status-only_b { color: #6a1b9a !important; }
}

/* 공통 */
.airs-header {
    display:flex; align-items:center; gap:16px;
    padding:20px 28px; margin-bottom:32px;
    border-radius:0 0 16px 16px;
}
.airs-logo-box {
    width:52px; height:52px; background:transparent;
    border-radius:12px;
    display:flex; align-items:center; justify-content:center; flex-shrink:0;
}
.airs-title h1 {
    margin:0; font-size:22px; font-weight:800;
    background:linear-gradient(90deg,#00d4ff,#0066ff);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent;
    letter-spacing:2px;
}
.airs-title p { margin:2px 0 0; font-size:13px; letter-spacing:1px; }
.airs-badge {
    margin-left:auto; padding:6px 14px;
    border-radius:20px; font-size:12px; font-weight:600; letter-spacing:1px;
}
.step-card { border-radius:16px; padding:20px 24px; margin-bottom:16px; }
.step-header { display:flex; align-items:center; gap:12px; }
.step-number {
    width:36px; height:36px;
    background:linear-gradient(135deg,#00d4ff,#0066ff);
    border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-weight:800; font-size:16px; color:white;
    box-shadow:0 2px 12px rgba(0,212,255,0.4); flex-shrink:0;
}
.step-title { font-size:18px; font-weight:700; margin:0; }
.success-banner {
    border-radius:12px; padding:14px 20px; font-weight:600;
    display:flex; align-items:center; gap:10px; margin:12px 0;
}
.sidebar-section-title {
    font-size:12px; font-weight:700;
    letter-spacing:1px; text-transform:uppercase;
    margin-bottom:10px; padding-bottom:6px;
}
.how-step { display:flex; align-items:flex-start; gap:10px; margin-bottom:8px; }
.how-step-num {
    width:20px; height:20px; flex-shrink:0;
    background:linear-gradient(135deg,#00d4ff,#0066ff);
    border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:11px; font-weight:700; color:white;
}
.how-step-text { font-size:13px; line-height:1.5; }
.mode-label {
    font-size:11px; font-weight:700; letter-spacing:1px; text-transform:uppercase;
    margin:12px 0 6px; display:flex; align-items:center; gap:6px;
}
.mode-label::after { content:''; flex:1; height:1px; }
.note-item { display:flex; align-items:flex-start; gap:8px; margin-bottom:8px; font-size:13px; }
.note-icon { font-size:14px; flex-shrink:0; }
.tag-row { display:flex; align-items:center; gap:6px; padding:5px 0; }
.tag-code { font-family:monospace; font-size:11px; padding:2px 6px; border-radius:4px; min-width:100px; flex-shrink:0; }
.tag-name { font-size:12px; flex:1; }
.tag-vr   { font-size:11px; padding:1px 5px; border-radius:3px; font-family:monospace; flex-shrink:0; }
.metric-card { border-radius:12px; padding:16px 20px; text-align:center; margin-bottom:8px; }
.metric-val  { font-size:28px; font-weight:800; }
.metric-lbl  { font-size:12px; font-weight:600; letter-spacing:1px; margin-top:4px; }
.file-header-a, .file-header-b {
    border-radius:10px; padding:10px 16px; margin-bottom:10px; font-weight:700; font-size:14px;
}
.diff-banner-warn, .diff-banner-ok {
    border-radius:12px; padding:16px 20px; margin-bottom:20px;
}
.apply-hint { border-radius:8px; padding:12px 16px; font-size:13px; font-weight:500; }
.dl-target-box { border-radius:14px; padding:20px 24px; margin:16px 0; }
.dl-target-title { font-size:18px; font-weight:800; margin-bottom:6px; }
.dl-target-desc  { font-size:13px; line-height:1.6; margin-top:8px; }
.table-info-text { font-size:13px; margin-bottom:6px; }
.page-ctrl-box {
    border-radius:10px; padding:8px 12px; margin-bottom:8px;
    display:flex; align-items:center; justify-content:center;
}
.page-num   { font-size:20px; font-weight:800; min-width:32px; text-align:center; }
.page-total { font-size:14px; font-weight:500; }
.manual-edit-box { border-radius:10px; padding:12px 16px; margin-bottom:12px; font-size:13px; }

/* ✅ Value Card */
.val-card {
    border-radius:10px; padding:12px 14px; margin-bottom:8px;
}
.val-label {
    font-size:10px; font-weight:700; letter-spacing:1px;
    text-transform:uppercase; margin-bottom:6px;
}
.val-content {
    font-family:monospace; font-size:11px;
    padding:7px 9px; border-radius:6px;
    word-break:break-all; line-height:1.5;
    max-height:80px; overflow-y:auto;
    white-space:pre-wrap;
}
.val-status-row {
    display:flex; align-items:center; gap:6px;
    margin-top:6px; font-size:11px; font-weight:600;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════
st.markdown(f"""
<div class="airs-header">
    <div class="airs-logo-box">{logo_html}</div>
    <div class="airs-title">
        <h1>SwiftMR</h1>
        <p>DICOM Header Editor &nbsp;·&nbsp; Internal Tool</p>
    </div>
    <div class="airs-badge">v2.0</div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# PAGE CONTROL
# ══════════════════════════════════════════════════════
def page_control(key: str, total_pages: int) -> int:
    cur_key = f"{key}_cur"
    if cur_key not in st.session_state:
        st.session_state[cur_key] = 1
    st.session_state[cur_key] = max(1, min(st.session_state[cur_key], total_pages))
    cur = st.session_state[cur_key]
    c1, c2, c3, c4, c5 = st.columns([1, 1, 4, 1, 1])
    with c1:
        if st.button("◀◀", key=f"{key}_first", use_container_width=True, disabled=(cur == 1)):
            st.session_state[cur_key] = 1; st.rerun()
    with c2:
        if st.button("◀", key=f"{key}_prev", use_container_width=True, disabled=(cur <= 1)):
            st.session_state[cur_key] -= 1; st.rerun()
    with c3:
        st.markdown(
            f'<div class="page-ctrl-box">'
            f'<span class="page-total" style="margin-right:8px;">Page</span>'
            f'<span class="page-num">{cur}</span>'
            f'<span class="page-total">&nbsp;/&nbsp;{total_pages}</span>'
            f'</div>', unsafe_allow_html=True)
    with c4:
        if st.button("▶", key=f"{key}_next", use_container_width=True, disabled=(cur >= total_pages)):
            st.session_state[cur_key] += 1; st.rerun()
    with c5:
        if st.button("▶▶", key=f"{key}_last", use_container_width=True, disabled=(cur == total_pages)):
            st.session_state[cur_key] = total_pages; st.rerun()
    return st.session_state[cur_key]

# ══════════════════════════════════════════════════════
# ✅ VALUE CARD — HTML 문자열 반환 (st.markdown 호출 없음)
# ══════════════════════════════════════════════════════
def val_card_html(label: str, value: str, status: str, show_status: bool = False) -> str:
    status_icons = {
        "match":  ("✅ Identical", "val-status-match"),
        "diff":   ("⚠️ Different", "val-status-diff"),
        "only_a": ("🔵 Only in A", "val-status-only_a"),
        "only_b": ("🟣 Only in B", "val-status-only_b"),
    }
    icon_text, icon_cls = status_icons.get(status, ("", ""))
    display_val = value if value else "— (empty)"
    safe_val    = html_mod.escape(display_val[:300]) + ("..." if len(display_val) > 300 else "")
    safe_label  = html_mod.escape(label)
    status_html = (
        f'<div class="val-status-row">'
        f'<span class="{icon_cls}">{icon_text}</span>'
        f'</div>'
        if show_status and icon_cls else ""
    )
    return (
        f'<div class="val-card">'
        f'<div class="val-label">{safe_label}</div>'
        f'<div class="val-content">{safe_val}</div>'
        f'{status_html}'
        f'</div>'
    )

# ══════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════
def parse_dicom(file_bytes: bytes):
    return pydicom.dcmread(io.BytesIO(file_bytes), force=True)

def is_dicom_bytes(data: bytes) -> bool:
    return len(data) >= 132 and data[128:132] == b"DICM"

def is_valid_dicom_ds(ds) -> bool:
    return len(ds) >= 3

def extract_tags(ds) -> list:
    rows = []
    for elem in ds:
        try:
            tag_str = f"({elem.tag.group:04X},{elem.tag.element:04X})"
            value   = (str(elem.value) if not isinstance(elem.value, bytes)
                       else f"[Binary {len(elem.value)} bytes]")
            rows.append({
                "Tag":     tag_str,
                "Keyword": elem.keyword if not elem.tag.is_private else "Private Tag",
                "VR":      str(elem.VR),
                "Value":   value,
                "Private": elem.tag.is_private,
            })
        except Exception:
            continue
    return rows

def apply_modifications_to_ds(ds, modifications: dict):
    ds_copy = deepcopy(ds)
    results = []
    for tag_str, new_value in modifications.items():
        try:
            group, element = tag_str.strip("()").split(",")
            tag = pydicom.tag.Tag(int(group, 16), int(element, 16))
            if tag in ds_copy:
                vr        = ds_copy[tag].VR
                old_value = str(ds_copy[tag].value)
                if vr in ["DS", "FL", "FD"]:
                    ds_copy[tag].value = float(new_value)
                elif vr in ["IS", "SL", "SS", "UL", "US"]:
                    ds_copy[tag].value = int(new_value)
                else:
                    ds_copy[tag].value = new_value
                results.append({
                    "Tag": tag_str, "Keyword": ds_copy[tag].keyword,
                    "Before": old_value, "After": new_value, "Status": "✅ Success",
                })
            else:
                results.append({
                    "Tag": tag_str, "Keyword": "-",
                    "Before": "-", "After": new_value, "Status": "⚠️ Not found",
                })
        except Exception as e:
            results.append({
                "Tag": tag_str, "Keyword": "-",
                "Before": "-", "After": new_value, "Status": f"❌ Error: {e}",
            })
    output = io.BytesIO()
    try:
        ds_copy.save_as(output, write_like_original=True)
    except Exception:
        try:
            if hasattr(ds_copy, 'file_meta'):
                pydicom.dataset.validate_file_meta(ds_copy.file_meta, enforce_standard=False)
            ds_copy.save_as(output, write_like_original=True)
        except Exception:
            ds_copy.is_implicit_VR   = False
            ds_copy.is_little_endian = True
            ds_copy.save_as(output, write_like_original=False)
    return output.getvalue(), results

def process_zip(zip_bytes: bytes, modifications: dict):
    input_zip  = zipfile.ZipFile(io.BytesIO(zip_bytes))
    output_buf = io.BytesIO()
    output_zip = zipfile.ZipFile(output_buf, "w", zipfile.ZIP_STORED)
    all_results = []
    success_count = skip_count = error_count = 0
    file_list = [f for f in input_zip.namelist() if not f.endswith("/")]
    skip_exts = {".jpg",".jpeg",".png",".gif",".bmp",".txt",".xml",".json",".pdf",".zip"}
    for filename in file_list:
        file_bytes = input_zip.read(filename)
        ext = Path(filename).suffix.lower()
        if ext in skip_exts:
            output_zip.writestr(filename, file_bytes); skip_count += 1; continue
        try:
            ds = pydicom.dcmread(io.BytesIO(file_bytes), force=True)
            if len(ds) < 3:
                output_zip.writestr(filename, file_bytes); skip_count += 1; continue
            modified_bytes, results = apply_modifications_to_ds(ds, modifications)
            for r in results: r["File"] = filename
            all_results.extend(results)
            output_zip.writestr(filename, modified_bytes)
            success_count += 1
        except Exception as e:
            error_count += 1
            all_results.append({"File": filename, "Tag": "-", "Keyword": "-",
                                 "Before": "-", "After": "-", "Status": f"❌ Failed: {e}"})
            output_zip.writestr(filename, file_bytes)
    output_zip.close()
    return output_buf.getvalue(), all_results, {
        "total": len(file_list), "success": success_count,
        "skipped": skip_count, "errors": error_count,
    }

def apply_staged_to_zip(zip_bytes: bytes, modifications: dict):
    output_buf = io.BytesIO()
    output_zip = zipfile.ZipFile(output_buf, "w", zipfile.ZIP_STORED)
    all_results = []
    success_count = skip_count = error_count = 0
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as input_zip:
        for filename in input_zip.namelist():
            if filename.endswith("/"):
                try: output_zip.writestr(filename, b"")
                except: pass
                continue
            file_bytes = input_zip.read(filename)
            base = Path(filename).name
            if base.startswith("._") or base.startswith("."):
                output_zip.writestr(filename, file_bytes); skip_count += 1; continue
            is_dcm = False
            try:
                ds = pydicom.dcmread(io.BytesIO(file_bytes), force=True)
                if is_valid_dicom_ds(ds): is_dcm = True
            except: pass
            if not is_dcm:
                output_zip.writestr(filename, file_bytes); skip_count += 1; continue
            try:
                modified_bytes, results = apply_modifications_to_ds(ds, modifications)
                for r in results: r["File"] = filename
                all_results.extend(results)
                output_zip.writestr(filename, modified_bytes)
                success_count += 1
            except Exception as e:
                error_count += 1
                all_results.append({"File": filename, "Tag": "-", "Keyword": "-",
                                     "Before": "-", "After": "-", "Status": f"❌ Failed: {e}"})
                output_zip.writestr(filename, file_bytes)
    output_zip.close()
    return output_buf.getvalue(), all_results, {
        "total": success_count + skip_count + error_count,
        "success": success_count, "skipped": skip_count, "errors": error_count,
    }

PROTECTED_TAGS = {"(7FE0,0010)"}

def get_dcm_list_from_zip(zip_bytes: bytes) -> list:
    result = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if name.endswith("/"): continue
            base = Path(name).name
            if base.startswith("._") or base.startswith("."): continue
            ext = Path(name).suffix.lower()
            if ext in {".jpg",".jpeg",".png",".gif",".bmp",".txt",".xml",".json",".pdf"}: continue
            data = zf.read(name)
            if ext in {".dcm",".dicom"} or is_dicom_bytes(data):
                try:
                    ds = pydicom.dcmread(io.BytesIO(data), force=True)
                    if is_valid_dicom_ds(ds): result.append(name)
                except: continue
    return sorted(result)

def read_dcm_from_zip(zip_bytes: bytes, filename: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        return zf.read(filename)

def compare_dicom(ds_a, ds_b) -> pd.DataFrame:
    def ds_to_dict(ds):
        d = {}
        for elem in ds:
            try:
                tag_str = f"({elem.tag.group:04X},{elem.tag.element:04X})"
                kw  = elem.keyword if not elem.tag.is_private else "Private Tag"
                vr  = str(elem.VR)
                val = (str(elem.value) if not isinstance(elem.value, bytes)
                       else f"[Binary {len(elem.value)} bytes]")
                d[tag_str] = {"Keyword": kw, "VR": vr, "Value": val}
            except: continue
        return d
    dict_a   = ds_to_dict(ds_a)
    dict_b   = ds_to_dict(ds_b)
    all_tags = sorted(set(dict_a) | set(dict_b))
    rows = []
    for tag in all_tags:
        in_a, in_b = tag in dict_a, tag in dict_b
        if in_a and in_b:
            kw, vr = dict_a[tag]["Keyword"], dict_a[tag]["VR"]
            va, vb = dict_a[tag]["Value"],   dict_b[tag]["Value"]
            status = "match" if va == vb else "diff"
        elif in_a:
            kw, vr, va, vb, status = (dict_a[tag]["Keyword"], dict_a[tag]["VR"],
                                       dict_a[tag]["Value"], "", "only_a")
        else:
            kw, vr, va, vb, status = (dict_b[tag]["Keyword"], dict_b[tag]["VR"],
                                       "", dict_b[tag]["Value"], "only_b")
        rows.append({"Tag": tag, "Keyword": kw, "VR": vr,
                     "Value A": va, "Value B": vb, "Status": status})
    return pd.DataFrame(rows)

def make_excel_diff(df: pd.DataFrame) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "DICOM Diff"
    fill_map = {
        "diff":   PatternFill("solid", fgColor="FFF3E0"),
        "only_a": PatternFill("solid", fgColor="E3F2FD"),
        "only_b": PatternFill("solid", fgColor="F3E5F5"),
        "match":  PatternFill("solid", fgColor="F1F8E9"),
    }
    ws.append(["Tag","Keyword","VR","Value A","Value B","Status"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CFD8DC")
    for _, row in df.iterrows():
        ws.append([row["Tag"], row["Keyword"], row["VR"],
                   row["Value A"], row["Value B"], row["Status"]])
        fill = fill_map.get(row["Status"], PatternFill())
        for cell in ws[ws.max_row]: cell.fill = fill
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════
defaults = {
    "ds": None, "tags_df": None, "modifications": {},
    "filename": "", "modified_bytes": None, "mod_results": None,
    "upload_mode": None, "zip_bytes": None, "summary": None, "queue_msg": None,
    "cmp_a_bytes": None, "cmp_a_name": "",
    "cmp_b_bytes": None, "cmp_b_name": "",
    "cmp_a_zip_bytes": None, "cmp_b_zip_bytes": None,
    "cmp_a_zip_list": [], "cmp_b_zip_list": [],
    "cmp_a_sel": "", "cmp_b_sel": "",
    "cmp_df": None,
    "cmp_staged": {},
    "cmp_manual_staged": {},
    "cmp_result_bytes": None, "cmp_result_zip": None,
    "cmp_result_summary": None, "cmp_result_log": None,
    "cmp_ds_b": None,
    "app_mode": "editor",
    "phi_confirmed_once": False,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ══════════════════════════════════════════════════════
# MODE SELECTOR
# ══════════════════════════════════════════════════════
col_m1, col_m2, col_m3 = st.columns([2, 1, 1])
with col_m2:
    if st.button("🔍 Validation Mode  Single / Batch DICOM tag validation",
                 use_container_width=True,
                 type="primary" if st.session_state.app_mode == "editor" else "secondary"):
        st.session_state.app_mode = "editor"; st.rerun()
with col_m3:
    if st.button("⚖️ Compare Mode  Side-by-side DICOM diff & inline edit",
                 use_container_width=True,
                 type="primary" if st.session_state.app_mode == "compare" else "secondary"):
        st.session_state.app_mode = "compare"; st.rerun()

st.markdown("<hr style='margin:8px 0 24px;opacity:0.2;'>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# PHI WARNING
# ══════════════════════════════════════════════════════
if not st.session_state.phi_confirmed_once:
    st.markdown("""
    <div style="background:linear-gradient(135deg,rgba(255,80,80,0.12),rgba(200,0,0,0.08));
        border:1.5px solid rgba(255,80,80,0.5); border-left:4px solid #ff4444;
        border-radius:12px; padding:16px 20px; margin-bottom:8px;">
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
            <span style="font-size:20px;">🔒</span>
            <span style="font-size:15px;font-weight:800;color:#ff4444;letter-spacing:1px;">
                HIPAA &amp; GDPR WARNING</span>
        </div>
        <div style="font-size:13px;line-height:1.8;">
            ⚠️ This tool runs on <b>Streamlit Cloud (external server)</b>.<br>
            ⚠️ <b>DO NOT upload files containing PHI (Protected Health Information)</b>.<br>
            ⚠️ Uploading real patient data may violate <b>HIPAA</b> and <b>GDPR</b> regulations.<br>
            ✅ Only use <b>fully anonymized or de-identified DICOM files</b>.
        </div>
    </div>
    """, unsafe_allow_html=True)
    if not st.checkbox("✅ I confirm that this file does NOT contain any PHI and is fully anonymized.",
                       key="phi_confirm"):
        st.warning("⛔ Please confirm the above statement before uploading any files.")
        st.stop()
    else:
        st.session_state.phi_confirmed_once = True; st.rerun()
else:
    st.markdown("""
    <div style="background:rgba(0,200,100,0.08);border:1px solid rgba(0,200,100,0.25);
        border-radius:8px;padding:8px 16px;font-size:12px;color:#00c864;margin-bottom:12px;
        display:flex;align-items:center;gap:8px;">
        🔒 <b>PHI Confirmed</b> — No real patient data will be uploaded.
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# ██████████  EDITOR MODE  ██████████
# ══════════════════════════════════════════════════════
if st.session_state.app_mode == "editor":

    st.markdown("""
    <div class="step-card" style="margin-bottom:24px;">
      <div class="step-header" style="gap:14px;">
        <span style="font-size:28px;">🔍</span>
        <div>
          <p class="step-title" style="margin-bottom:4px;">Validation Mode — Single / Batch DICOM Tag Editor</p>
          <p style="margin:0;font-size:13px;color:#8892a4;">
            Upload a single .dcm or a .zip archive &nbsp;·&nbsp;
            Search &amp; edit tags &nbsp;·&nbsp; Download modified file
          </p>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    st.markdown("""<div class="step-card"><div class="step-header">
        <div class="step-number">1</div><p class="step-title">Upload File</p>
    </div></div>""", unsafe_allow_html=True)

    upload_mode = st.radio("mode", ["🗂️ Single DICOM (.dcm)", "📦 Multiple DICOMs (.zip)"],
                           horizontal=True, label_visibility="collapsed")

    if "Single" in upload_mode:
        uploaded = st.file_uploader("Upload a DICOM file", type=["dcm","DCM"])
        if uploaded:
            if uploaded.name != st.session_state.filename:
                fb = uploaded.read()
                st.session_state.ds            = parse_dicom(fb)
                st.session_state.tags_df       = extract_tags(st.session_state.ds)
                st.session_state.filename      = uploaded.name
                st.session_state.upload_mode   = "single"
                st.session_state.zip_bytes     = None
                st.session_state.modifications = {}
                st.session_state.modified_bytes= None
                st.session_state.mod_results   = None
                st.session_state.summary       = None
                st.session_state.queue_msg     = None
                st.session_state["editor_page_cur"] = 1
            n_tags = len(st.session_state.tags_df) if st.session_state.tags_df else 0
            st.markdown(f"""
            <div class="diff-banner-ok">
                <div style="font-size:16px;font-weight:800;margin-bottom:4px;">✅ File Loaded Successfully</div>
                <div style="font-size:13px;color:#8892a4;">
                    <b>{st.session_state.filename}</b> &nbsp;·&nbsp; {n_tags} tags found
                </div>
            </div>""", unsafe_allow_html=True)
    else:
        uploaded = st.file_uploader("Upload a ZIP file", type=["zip"])
        if uploaded:
            if uploaded.name != st.session_state.filename:
                zb = uploaded.read()
                first_dcm_bytes = None
                with zipfile.ZipFile(io.BytesIO(zb)) as zf:
                    for fname in zf.namelist():
                        if fname.endswith("/"): continue
                        fb = zf.read(fname)
                        try:
                            ds_test = pydicom.dcmread(io.BytesIO(fb), force=True)
                            if len(ds_test) >= 3: first_dcm_bytes = fb; break
                        except: continue
                if first_dcm_bytes:
                    st.session_state.ds            = parse_dicom(first_dcm_bytes)
                    st.session_state.tags_df       = extract_tags(st.session_state.ds)
                    st.session_state.zip_bytes     = zb
                    st.session_state.filename      = uploaded.name
                    st.session_state.upload_mode   = "zip"
                    st.session_state.modifications = {}
                    st.session_state.modified_bytes= None
                    st.session_state.mod_results   = None
                    st.session_state.summary       = None
                    st.session_state.queue_msg     = None
                    st.session_state["editor_page_cur"] = 1
                else:
                    st.error("❌ No valid DICOM files found in ZIP.")
            if st.session_state.filename and st.session_state.tags_df:
                with zipfile.ZipFile(io.BytesIO(st.session_state.zip_bytes)) as zf:
                    file_list = [f for f in zf.namelist() if not f.endswith("/")]
                st.markdown(f"""
                <div class="diff-banner-ok">
                    <div style="font-size:16px;font-weight:800;margin-bottom:4px;">✅ ZIP Loaded Successfully</div>
                    <div style="font-size:13px;color:#8892a4;">
                        <b>{st.session_state.filename}</b>
                        &nbsp;·&nbsp; {len(file_list)} files
                        &nbsp;·&nbsp; {len(st.session_state.tags_df)} tags
                    </div>
                </div>""", unsafe_allow_html=True)
                with st.expander(f"📂 View {len(file_list)} files in ZIP", expanded=False):
                    st.dataframe(pd.DataFrame({"File": file_list}),
                                 hide_index=True, use_container_width=True)

    if st.session_state.ds is not None:
        st.markdown("""<div class="step-card"><div class="step-header">
            <div class="step-number">2</div><p class="step-title">View &amp; Search Tags</p>
        </div></div>""", unsafe_allow_html=True)

        df_all    = pd.DataFrame(st.session_state.tags_df)
        n_total   = len(df_all)
        n_private = int(df_all["Private"].sum())
        n_std     = n_total - n_private

        mc = st.columns(3)
        for col, (lbl, val, color) in zip(mc, [
            ("TOTAL TAGS",   n_total,   "#607d8b"),
            ("STANDARD",     n_std,     "#2196f3"),
            ("PRIVATE TAGS", n_private, "#9c27b0"),
        ]):
            col.markdown(
                f'<div class="metric-card">'
                f'<div class="metric-val" style="color:{color};">{val}</div>'
                f'<div class="metric-lbl">{lbl}</div></div>',
                unsafe_allow_html=True)

        st.markdown("### 🔎 Filter & Search")
        fs1, fs2, fs3 = st.columns([3, 2, 1])
        with fs1:
            search = st.text_input("Search", placeholder="e.g. PatientName, 0010,0010",
                                   key="editor_search", label_visibility="collapsed")
        with fs2:
            show_private = st.selectbox("Tag type",
                                        ["All Tags","Standard Only","Private Only"],
                                        key="editor_tag_type", label_visibility="collapsed")
        with fs3:
            ed_page_size = st.selectbox("Rows", [25,50,100,200],
                                        key="editor_page_size", label_visibility="collapsed")

        df = df_all.copy()
        if show_private == "Standard Only": df = df[df["Private"] == False]
        elif show_private == "Private Only": df = df[df["Private"] == True]
        if search:
            esc  = re.escape(search)
            mask = (df["Keyword"].str.contains(esc, case=False, na=False, regex=True)
                    | df["Tag"].str.contains(esc, case=False, na=False, regex=True)
                    | df["Value"].str.contains(esc, case=False, na=False, regex=True))
            df = df[mask]

        ed_total_pages = max(1, (len(df) - 1) // ed_page_size + 1)
        st.markdown("### 🗂️ Tag Table")
        st.markdown(
            f'<div class="table-info-text">Showing <b>{min(ed_page_size, len(df))}</b> of '
            f'<b>{len(df)}</b> tags &nbsp;·&nbsp; total: {n_total}</div>',
            unsafe_allow_html=True)
        ed_page    = page_control("editor_page", ed_total_pages)
        df_page_ed = df.iloc[(ed_page-1)*ed_page_size : ed_page*ed_page_size].copy()

        def highlight_modified_editor(row):
            if row["Tag"] in st.session_state.modifications:
                return ["background-color:rgba(0,212,255,0.15)"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df_page_ed.drop(columns=["Private"]).style.apply(highlight_modified_editor, axis=1),
            use_container_width=True, height=420, hide_index=True)

        st.markdown("""<div class="step-card"><div class="step-header">
            <div class="step-number">3</div><p class="step-title">Edit Tags</p>
        </div></div>""", unsafe_allow_html=True)
        st.caption("Select a tag from the filtered list, enter a new value, and click **Queue Change**.")

        col_a, col_b = st.columns([2, 3])
        with col_a:
            tag_options     = [f"{row['Tag']}  {row['Keyword']}" for _, row in df.iterrows()]
            selected_option = st.selectbox("Select tag to edit", tag_options, key="tag_select")
            selected_tag    = selected_option.split("  ")[0] if selected_option else None
        with col_b:
            if selected_tag:
                cur         = df[df["Tag"] == selected_tag]["Value"].values
                default_val = st.session_state.modifications.get(
                    selected_tag, cur[0] if len(cur) else "")
            else:
                default_val = ""
            new_value = st.text_input("New value", value=default_val)

        if st.button("📝 Queue Change", use_container_width=True, key="queue_btn"):
            val = new_value.strip()
            if not selected_tag:
                st.session_state.queue_msg = ("error", "Please select a tag.")
            elif not val:
                st.session_state.queue_msg = ("error", "Please enter a new value.")
            else:
                st.session_state.modifications[selected_tag] = val
                st.session_state.modified_bytes = None
                st.session_state.mod_results    = None
                st.session_state.summary        = None
                st.session_state.queue_msg = ("success", f"Queued: **{selected_tag}** → `{val}`")

        if st.session_state.queue_msg:
            t, m = st.session_state.queue_msg
            st.success(m) if t == "success" else st.warning(m)

        if st.session_state.modifications:
            st.markdown("#### 📋 Pending Modifications")
            atdf = pd.DataFrame(st.session_state.tags_df)
            mod_rows = []
            for tag, val in st.session_state.modifications.items():
                kw  = atdf[atdf["Tag"] == tag]["Keyword"].values
                ori = atdf[atdf["Tag"] == tag]["Value"].values
                mod_rows.append({
                    "Tag": tag,
                    "Keyword":  kw[0]  if len(kw)  else "Unknown",
                    "Original": ori[0] if len(ori) else "-",
                    "New Value": val,
                })
            st.dataframe(pd.DataFrame(mod_rows), use_container_width=True, hide_index=True)
            c1, c2 = st.columns([2, 3])
            with c1:
                if st.button("🗑️ Clear All", use_container_width=True):
                    st.session_state.modifications  = {}
                    st.session_state.modified_bytes = None
                    st.session_state.mod_results    = None
                    st.session_state.summary        = None
                    st.session_state.queue_msg      = None
                    st.rerun()
            with c2:
                tag_to_remove = st.selectbox("Remove specific tag",
                                             list(st.session_state.modifications.keys()),
                                             key="remove_select")
                if st.button("❌ Remove Selected", use_container_width=True):
                    del st.session_state.modifications[tag_to_remove]
                    st.session_state.modified_bytes = None
                    st.session_state.mod_results    = None
                    st.session_state.queue_msg      = None
                    st.rerun()

        st.markdown("""<div class="step-card"><div class="step-header">
            <div class="step-number">4</div><p class="step-title">Apply &amp; Download</p>
        </div></div>""", unsafe_allow_html=True)

        if not st.session_state.modifications:
            st.markdown(
                '<div class="apply-hint">💡 Queue at least one tag change above to enable download.</div>',
                unsafe_allow_html=True)
        else:
            n_mods = len(st.session_state.modifications)
            if st.session_state.upload_mode == "single":
                st.markdown(f"""<div class="dl-target-box">
                    <div class="dl-target-title">📄 Download Target</div>
                    <div class="dl-target-desc">
                        <b>Mode:</b> Single DICOM file<br>
                        <b>File:</b> {st.session_state.filename}<br>
                        <b>Changes:</b> {n_mods} tag(s) will be modified<br>
                        <b>Output:</b> modified_{st.session_state.filename}
                    </div></div>""", unsafe_allow_html=True)
                if st.button("🚀 Apply Changes", type="primary", use_container_width=True):
                    with st.spinner("Processing..."):
                        mb, results = apply_modifications_to_ds(
                            st.session_state.ds, st.session_state.modifications)
                    st.session_state.modified_bytes = mb
                    st.session_state.mod_results    = results
                if st.session_state.mod_results:
                    st.dataframe(pd.DataFrame(st.session_state.mod_results),
                                 use_container_width=True, hide_index=True)
                if st.session_state.modified_bytes:
                    st.markdown('<div class="success-banner">✅ File ready for download!</div>',
                                unsafe_allow_html=True)
                    st.download_button(
                        label="⬇️ Download Modified DICOM",
                        data=st.session_state.modified_bytes,
                        file_name=f"modified_{st.session_state.filename}",
                        mime="application/octet-stream",
                        use_container_width=True, type="primary")
            elif st.session_state.upload_mode == "zip":
                with zipfile.ZipFile(io.BytesIO(st.session_state.zip_bytes)) as zf:
                    n_files = len([f for f in zf.namelist() if not f.endswith("/")])
                zip_out = st.session_state.filename.replace(".zip", "_modified.zip")
                st.markdown(f"""<div class="dl-target-box">
                    <div class="dl-target-title">📦 Download Target</div>
                    <div class="dl-target-desc">
                        <b>Mode:</b> Batch ZIP (no compression — DICOM safe)<br>
                        <b>File:</b> {st.session_state.filename} &nbsp;·&nbsp; {n_files} files<br>
                        <b>Changes:</b> {n_mods} tag(s) applied to every DICOM<br>
                        <b>Output:</b> {zip_out}
                    </div></div>""", unsafe_allow_html=True)
                if st.button("🚀 Apply to All & Create ZIP", type="primary", use_container_width=True):
                    with st.spinner("Processing ZIP..."):
                        rb, all_results, summary = process_zip(
                            st.session_state.zip_bytes, st.session_state.modifications)
                    st.session_state.modified_bytes = rb
                    st.session_state.mod_results    = all_results
                    st.session_state.summary        = summary
                if st.session_state.summary is not None:
                    s = st.session_state.summary
                    mc1, mc2, mc3, mc4 = st.columns(4)
                    mc1.metric("📁 Total",     s["total"])
                    mc2.metric("✅ Processed", s["success"])
                    mc3.metric("⏭️ Skipped",   s["skipped"])
                    mc4.metric("❌ Errors",    s["errors"])
                    if s["success"] == 0: st.error("❌ No DICOM files were processed!")
                if st.session_state.mod_results:
                    with st.expander("📊 Modification Report", expanded=False):
                        st.dataframe(pd.DataFrame(st.session_state.mod_results),
                                     use_container_width=True, height=400, hide_index=True)
                if st.session_state.modified_bytes is not None:
                    st.markdown('<div class="success-banner">✅ ZIP ready for download!</div>',
                                unsafe_allow_html=True)
                    st.download_button(
                        label="⬇️ Download Modified ZIP",
                        data=st.session_state.modified_bytes,
                        file_name=zip_out, mime="application/zip",
                        use_container_width=True, type="primary")

# ══════════════════════════════════════════════════════
# ██████████  COMPARE MODE  ██████████
# ══════════════════════════════════════════════════════
else:
    st.markdown("""
    <div class="step-card" style="margin-bottom:24px;">
      <div class="step-header" style="gap:14px;">
        <span style="font-size:28px;">⚖️</span>
        <div>
          <p class="step-title" style="margin-bottom:4px;">Compare Mode — Side-by-Side DICOM Diff</p>
          <p style="margin:0;font-size:13px;color:#8892a4;">
            Upload two DICOM files (or ZIP archives) to compare all tags
            &nbsp;·&nbsp; Highlighted differences &nbsp;·&nbsp; Inline value editing
          </p>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown('<div class="file-header-a">📄 &nbsp;<b style="color:#00d4ff;">File A</b>'
                    ' &nbsp; Reference / Source</div>', unsafe_allow_html=True)
        st.caption("Upload File A (.dcm or .zip)")
        up_a = st.file_uploader("Upload File A", type=["dcm","DCM","zip"],
                                key="cmp_up_a", label_visibility="collapsed")
        if up_a:
            raw = up_a.read()
            if up_a.name.lower().endswith(".zip"):
                if up_a.name != st.session_state.cmp_a_name:
                    with st.spinner("Reading ZIP A..."):
                        dcm_list = get_dcm_list_from_zip(raw)
                    st.session_state.cmp_a_zip_bytes   = raw
                    st.session_state.cmp_a_zip_list    = dcm_list
                    st.session_state.cmp_a_bytes       = None
                    st.session_state.cmp_a_name        = up_a.name
                    st.session_state.cmp_a_sel         = ""
                    st.session_state.cmp_df            = None
                    st.session_state.cmp_staged        = {}
                    st.session_state.cmp_manual_staged = {}
                    st.session_state["cmp_page_cur"]   = 1
            else:
                if up_a.name != st.session_state.cmp_a_name:
                    st.session_state.cmp_a_bytes       = raw
                    st.session_state.cmp_a_name        = up_a.name
                    st.session_state.cmp_a_zip_bytes   = None
                    st.session_state.cmp_a_zip_list    = []
                    st.session_state.cmp_a_sel         = ""
                    st.session_state.cmp_df            = None
                    st.session_state.cmp_staged        = {}
                    st.session_state.cmp_manual_staged = {}
                    st.session_state["cmp_page_cur"]   = 1
        if st.session_state.cmp_a_zip_list:
            st.success(f"✅ File A ZIP: {len(st.session_state.cmp_a_zip_list)} DICOM files found")
            name_map_a    = {Path(p).name: p for p in st.session_state.cmp_a_zip_list}
            sel_a_display = st.selectbox("📄 Select DICOM from File A ZIP",
                                         list(name_map_a.keys()), key="cmp_a_sel_box")
            sel_a = name_map_a[sel_a_display]
            if sel_a != st.session_state.cmp_a_sel:
                st.session_state.cmp_a_sel         = sel_a
                st.session_state.cmp_a_bytes       = read_dcm_from_zip(
                    st.session_state.cmp_a_zip_bytes, sel_a)
                st.session_state.cmp_df            = None
                st.session_state.cmp_staged        = {}
                st.session_state.cmp_manual_staged = {}
                st.session_state["cmp_page_cur"]   = 1
        elif st.session_state.cmp_a_bytes:
            st.success(f"✅ A: **{st.session_state.cmp_a_name}**")

    with col_b:
        st.markdown('<div class="file-header-b">📄 &nbsp;<b style="color:#9c27b0;">File B</b>'
                    ' &nbsp; Target / Modified</div>', unsafe_allow_html=True)
        st.caption("Upload File B (.dcm or .zip)")
        up_b = st.file_uploader("Upload File B", type=["dcm","DCM","zip"],
                                key="cmp_up_b", label_visibility="collapsed")
        if up_b:
            raw = up_b.read()
            if up_b.name.lower().endswith(".zip"):
                if up_b.name != st.session_state.cmp_b_name:
                    with st.spinner("Reading ZIP B..."):
                        dcm_list = get_dcm_list_from_zip(raw)
                    st.session_state.cmp_b_zip_bytes   = raw
                    st.session_state.cmp_b_zip_list    = dcm_list
                    st.session_state.cmp_b_bytes       = None
                    st.session_state.cmp_b_name        = up_b.name
                    st.session_state.cmp_b_sel         = ""
                    st.session_state.cmp_df            = None
                    st.session_state.cmp_staged        = {}
                    st.session_state.cmp_manual_staged = {}
                    st.session_state["cmp_page_cur"]   = 1
            else:
                if up_b.name != st.session_state.cmp_b_name:
                    st.session_state.cmp_b_bytes       = raw
                    st.session_state.cmp_b_name        = up_b.name
                    st.session_state.cmp_b_zip_bytes   = None
                    st.session_state.cmp_b_zip_list    = []
                    st.session_state.cmp_b_sel         = ""
                    st.session_state.cmp_df            = None
                    st.session_state.cmp_staged        = {}
                    st.session_state.cmp_manual_staged = {}
                    st.session_state["cmp_page_cur"]   = 1
        if st.session_state.cmp_b_zip_list:
            st.success(f"✅ File B ZIP: {len(st.session_state.cmp_b_zip_list)} DICOM files found")
            name_map_b    = {Path(p).name: p for p in st.session_state.cmp_b_zip_list}
            sel_b_display = st.selectbox("📄 Select DICOM from File B ZIP",
                                         list(name_map_b.keys()), key="cmp_b_sel_box")
            sel_b = name_map_b[sel_b_display]
            if sel_b != st.session_state.cmp_b_sel:
                st.session_state.cmp_b_sel         = sel_b
                st.session_state.cmp_b_bytes       = read_dcm_from_zip(
                    st.session_state.cmp_b_zip_bytes, sel_b)
                st.session_state.cmp_df            = None
                st.session_state.cmp_staged        = {}
                st.session_state.cmp_manual_staged = {}
                st.session_state["cmp_page_cur"]   = 1
        elif st.session_state.cmp_b_bytes:
            st.success(f"✅ B: **{st.session_state.cmp_b_name}**")

    can_compare = (st.session_state.cmp_a_bytes is not None
                   and st.session_state.cmp_b_bytes is not None)
    st.markdown("<br>", unsafe_allow_html=True)

    if can_compare:
        if st.button("🔍 Run Comparison", type="primary", use_container_width=True):
            with st.spinner("Comparing DICOM headers..."):
                ds_a = parse_dicom(st.session_state.cmp_a_bytes)
                ds_b = parse_dicom(st.session_state.cmp_b_bytes)
                st.session_state.cmp_df             = compare_dicom(ds_a, ds_b)
                st.session_state.cmp_ds_b           = ds_b
                st.session_state.cmp_staged         = {}
                st.session_state.cmp_manual_staged  = {}
                st.session_state.cmp_result_bytes   = None
                st.session_state.cmp_result_zip     = None
                st.session_state.cmp_result_summary = None
                st.session_state.cmp_result_log     = None
                st.session_state["cmp_page_cur"]    = 1
    else:
        st.info("⬆️ Upload both File A and File B to enable comparison.")

    if st.session_state.cmp_df is not None:
        df_full  = st.session_state.cmp_df
        n_total  = len(df_full)
        n_match  = (df_full["Status"] == "match").sum()
        n_diff   = (df_full["Status"] == "diff").sum()
        n_only_a = (df_full["Status"] == "only_a").sum()
        n_only_b = (df_full["Status"] == "only_b").sum()
        n_issues = n_diff + n_only_a + n_only_b

        st.markdown("<br>", unsafe_allow_html=True)

        if n_issues == 0:
            st.markdown(f"""<div class="diff-banner-ok">
                <div style="font-size:20px;font-weight:800;margin-bottom:4px;">✅ FILES ARE IDENTICAL</div>
                <div style="font-size:13px;color:#8892a4;">All {n_total} tags match perfectly.</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class="diff-banner-warn">
                <div style="font-size:20px;font-weight:800;margin-bottom:4px;">
                    ⚠️ {n_issues} DIFFERENCE(S) FOUND</div>
                <div style="font-size:13px;color:#8892a4;">
                    Value differences: {n_diff} &nbsp;·&nbsp;
                    Only in A: {n_only_a} &nbsp;·&nbsp; Only in B: {n_only_b}
                </div>
            </div>""", unsafe_allow_html=True)

        mc = st.columns(5)
        for col, (lbl, val, color) in zip(mc, [
            ("TOTAL TAGS",   n_total,  "#607d8b"),
            ("✅ IDENTICAL", n_match,  "#4caf50"),
            ("⚠️ DIFFERENT", n_diff,   "#ff9800"),
            ("🔵 ONLY A",    n_only_a, "#2196f3"),
            ("🟣 ONLY B",    n_only_b, "#9c27b0"),
        ]):
            col.markdown(
                f'<div class="metric-card">'
                f'<div class="metric-val" style="color:{color};">{val}</div>'
                f'<div class="metric-lbl">{lbl}</div></div>',
                unsafe_allow_html=True)

        st.markdown("""<div style="display:flex;gap:20px;flex-wrap:wrap;
            margin:12px 0 20px;font-size:12px;">
            <span>✅ <b>Identical</b></span>
            <span>⚠️ <b>Different</b></span>
            <span>🔵 <b>Only in A</b></span>
            <span>🟣 <b>Only in B</b></span>
        </div>""", unsafe_allow_html=True)

        STATUS_COLORS = {
            "diff":   "rgba(255,152,0,0.15)",
            "only_a": "rgba(33,150,243,0.12)",
            "only_b": "rgba(156,39,176,0.12)",
            "match":  "rgba(76,175,80,0.08)",
        }
        STATUS_ICONS = {
            "diff":   "⚠️ Different",
            "only_a": "🔵 Only A",
            "only_b": "🟣 Only B",
            "match":  "✅ match",
        }

        st.markdown("### 🔎 Filter & Search")
        fs1, fs2, fs3 = st.columns([3, 2, 1])
        with fs1:
            cmp_search = st.text_input("Search", placeholder="e.g. PatientID, 0010,0020",
                                       key="cmp_search", label_visibility="collapsed")
        with fs2:
            show_opt = st.selectbox("Show",
                                    ["All","Different Only","Only in A","Only in B","Identical"],
                                    key="cmp_show_opt", label_visibility="collapsed")
        with fs3:
            page_size = st.selectbox("Rows per page", [25,50,100,200],
                                     key="cmp_page_size", label_visibility="collapsed")

        status_map = {
            "All":           ["diff","only_a","only_b","match"],
            "Different Only":["diff"],
            "Only in A":     ["only_a"],
            "Only in B":     ["only_b"],
            "Identical":     ["match"],
        }
        df_view = df_full[df_full["Status"].isin(status_map[show_opt])].copy()
        if cmp_search:
            esc  = re.escape(cmp_search)
            mask = (df_view["Tag"].str.contains(esc, case=False, na=False)
                    | df_view["Keyword"].str.contains(esc, case=False, na=False)
                    | df_view["Value A"].str.contains(esc, case=False, na=False)
                    | df_view["Value B"].str.contains(esc, case=False, na=False))
            df_view = df_view[mask]

        total_pages = max(1, (len(df_view) - 1) // page_size + 1)

        st.markdown("### 🗂️ Comparison Table")
        st.caption("✏️ Expand a tag in the sections below to edit its value in File B.")
        st.markdown(
            f'<div class="table-info-text">Showing <b>{min(page_size, len(df_view))}</b> of '
            f'<b>{len(df_view)}</b> filtered rows &nbsp;·&nbsp; total: {n_total}</div>',
            unsafe_allow_html=True)
        page    = page_control("cmp_page", total_pages)
        df_page = df_view.iloc[(page-1)*page_size : page*page_size].copy()

        df_display         = df_page[["Tag","Keyword","VR","Value A","Value B","Status"]].copy()
        df_display["Diff"] = df_display["Status"].map(STATUS_ICONS)

        def color_rows(row):
            c = STATUS_COLORS.get(row["Status"], "")
            return [f"background-color:{c}"] * len(row)

        styled = df_display.style.apply(color_rows, axis=1).hide(subset=["Status"], axis="columns")
        st.dataframe(styled, use_container_width=True, height=420, hide_index=True)

        # ── Inline Edit ───────────────────────────────
        st.markdown("---")
        st.markdown("### ✏️ Inline Edit — Apply Changes to File B")

        edit_filter = st.radio("Show tags",
                               ["⚠️ Different", "🔵 Only in A", "🟣 Only in B", "📋 All"],
                               horizontal=True, key="inline_edit_filter")
        edit_status_map = {
            "⚠️ Different": ["diff"],
            "🔵 Only in A":  ["only_a"],
            "🟣 Only in B":  ["only_b"],
            "📋 All":        ["diff","only_a","only_b"],
        }
        editable_df = df_full[
            df_full["Status"].isin(edit_status_map[edit_filter])
            & ~df_full["Tag"].isin(PROTECTED_TAGS)
        ].copy()

        st.caption(f"**{len(editable_df)}** editable tag(s) · "
                   "Stage changes then click Apply & Download.")

        if editable_df.empty:
            st.info("🎉 No differences — files are identical!" if edit_filter == "📋 All"
                    else f"No tags found for filter: **{edit_filter}**")
        else:
            for _, row in editable_df.iterrows():
                tag       = row["Tag"]
                kw        = row["Keyword"]
                status    = row["Status"]
                icon      = STATUS_ICONS.get(status, "")
                is_staged = tag in st.session_state.cmp_staged
                label     = f"{icon} | {tag}  {kw}" + (" ✅ Staged" if is_staged else "")

                with st.expander(label, expanded=False):
                    e1, e2 = st.columns([3, 2])
                    with e1:
                        # ✅ 두 카드를 합쳐서 한 번에 렌더링
                        st.markdown(
                            val_card_html("🔵 VALUE A — REFERENCE", row["Value A"], status)
                            + val_card_html("🟣 VALUE B — CURRENT", row["Value B"], status,
                                            show_status=True),
                            unsafe_allow_html=True
                        )
                    with e2:
                        st.markdown("**New value for B**")
                        current_staged = st.session_state.cmp_staged.get(tag, row["Value B"])
                        new_val = st.text_area("New value for B", value=current_staged,
                                               key=f"edit_{tag}", height=80,
                                               label_visibility="collapsed")
                        if st.button(f"📌 Stage{' ✅' if is_staged else ''}",
                                     key=f"stage_{tag}", use_container_width=True):
                            v = new_val.strip()
                            if v:
                                st.session_state.cmp_staged[tag] = v
                                st.session_state.cmp_result_bytes   = None
                                st.session_state.cmp_result_zip     = None
                                st.session_state.cmp_result_summary = None
                                st.rerun()
                            else:
                                st.warning("Please enter a value.")

        # ── Manual Edit ───────────────────────────────
        st.markdown("---")
        st.markdown("### 🔧 Manual Edit — Search & Edit Any Tag in File B")
        st.markdown(
            '<div class="manual-edit-box">'
            '🔍 &nbsp;Search <b>any tag</b> (including identical ones) '
            'and edit its value in File B. '
            'Staged changes are applied together with Inline Edits on download.'
            '</div>', unsafe_allow_html=True)

        manual_search = st.text_input(
            "🔍 Search tag",
            placeholder="e.g. SeriesDescription, 0008,103E",
            key="manual_search", label_visibility="collapsed")

        all_tags_df = df_full[~df_full["Tag"].isin(PROTECTED_TAGS)].copy()

        if manual_search:
            esc  = re.escape(manual_search)
            mask = (all_tags_df["Tag"].str.contains(esc, case=False, na=False)
                    | all_tags_df["Keyword"].str.contains(esc, case=False, na=False)
                    | all_tags_df["Value B"].str.contains(esc, case=False, na=False)
                    | all_tags_df["Value A"].str.contains(esc, case=False, na=False))
            search_result_df = all_tags_df[mask].copy()
        else:
            search_result_df = pd.DataFrame()

        if manual_search and search_result_df.empty:
            st.warning("⚠️ No matching tags found.")
        elif not search_result_df.empty:
            st.caption(f"**{len(search_result_df)}** tag(s) found")
            for _, row in search_result_df.iterrows():
                tag       = row["Tag"]
                kw        = row["Keyword"]
                status    = row["Status"]
                icon      = STATUS_ICONS.get(status, "🔵")
                is_staged = tag in st.session_state.cmp_manual_staged
                label     = f"{icon} | {tag}  {kw}" + (" ✅ Staged" if is_staged else "")

                with st.expander(label, expanded=True):
                    m1, m2 = st.columns([3, 2])
                    with m1:
                        # ✅ 두 카드를 합쳐서 한 번에 렌더링
                        st.markdown(
                            val_card_html("🔵 VALUE A — REFERENCE", row["Value A"], status)
                            + val_card_html("🟣 VALUE B — CURRENT", row["Value B"], status,
                                            show_status=True),
                            unsafe_allow_html=True
                        )
                    with m2:
                        st.markdown("**New value for B**")
                        cur_manual = st.session_state.cmp_manual_staged.get(tag, row["Value B"])
                        manual_val = st.text_area("New value for B", value=cur_manual,
                                                  key=f"manual_edit_{tag}", height=80,
                                                  label_visibility="collapsed")
                        mb1, mb2 = st.columns(2)
                        with mb1:
                            if st.button("📌 Stage" if not is_staged else "📌 Update",
                                         key=f"manual_stage_{tag}",
                                         use_container_width=True, type="primary"):
                                v = manual_val.strip()
                                if v:
                                    st.session_state.cmp_manual_staged[tag] = v
                                    st.session_state.cmp_result_bytes   = None
                                    st.session_state.cmp_result_zip     = None
                                    st.session_state.cmp_result_summary = None
                                    st.rerun()
                                else:
                                    st.warning("Please enter a value.")
                        with mb2:
                            if is_staged:
                                if st.button("🗑️ Remove", key=f"manual_remove_{tag}",
                                             use_container_width=True):
                                    del st.session_state.cmp_manual_staged[tag]
                                    st.rerun()

        if st.session_state.cmp_manual_staged:
            st.markdown(f"**📋 Manual Staged: {len(st.session_state.cmp_manual_staged)} tag(s)**")
            manual_rows = []
            for tag, val in st.session_state.cmp_manual_staged.items():
                orig = df_full[df_full["Tag"] == tag]["Value B"].values
                manual_rows.append({
                    "Tag": tag,
                    "Original B": orig[0] if len(orig) else "-",
                    "New Value":  val,
                })
            st.dataframe(pd.DataFrame(manual_rows), use_container_width=True, hide_index=True)
            if st.button("🗑️ Clear All Manual Staged", use_container_width=True):
                st.session_state.cmp_manual_staged = {}
                st.rerun()

        # ── Apply & Download ──────────────────────────
        st.markdown("---")
        st.markdown("### 💾 Apply & Download Modified File B")

        merged_staged = {
            **st.session_state.cmp_staged,
            **st.session_state.cmp_manual_staged,
        }

        if not merged_staged:
            st.markdown(
                '<div class="apply-hint">💡 Stage at least one edit above to enable download.</div>',
                unsafe_allow_html=True)
        else:
            staged_rows = []
            for tag, val in merged_staged.items():
                orig   = df_full[df_full["Tag"] == tag]["Value B"].values
                source = ("🔧 Manual" if tag in st.session_state.cmp_manual_staged
                          else "✏️ Inline")
                staged_rows.append({
                    "Source":     source,
                    "Tag":        tag,
                    "Original B": orig[0] if len(orig) else "-",
                    "New Value":  val,
                })
            st.dataframe(pd.DataFrame(staged_rows), use_container_width=True, hide_index=True)

            has_zip_b = st.session_state.cmp_b_zip_bytes is not None
            if has_zip_b:
                dl_mode = st.radio(
                    "dl_mode_radio",
                    ["📄 Single file — selected DCM only",
                     "📦 Full ZIP — apply to ALL files in ZIP B"],
                    key="cmp_dl_mode", horizontal=True, label_visibility="collapsed")
                is_zip_mode = "Full ZIP" in dl_mode
            else:
                is_zip_mode = False

            n_staged = len(merged_staged)

            if is_zip_mode:
                n_zip_files  = len(st.session_state.cmp_b_zip_list)
                zip_out_name = st.session_state.cmp_b_name.replace(".zip", "_modified.zip")
                st.markdown(f"""<div class="dl-target-box">
                    <div class="dl-target-title">📦 Download Target — Full ZIP</div>
                    <div class="dl-target-desc">
                        <b>Mode:</b> Batch — ALL {n_zip_files} DICOM files (no compression)<br>
                        <b>Source ZIP:</b> {st.session_state.cmp_b_name}<br>
                        <b>Changes:</b> {n_staged} tag(s) applied to every DICOM<br>
                        <b>Output:</b> {zip_out_name}
                    </div></div>""", unsafe_allow_html=True)
            else:
                b_base = Path(st.session_state.cmp_b_name).stem
                b_sel  = st.session_state.cmp_b_sel or st.session_state.cmp_b_name
                st.markdown(f"""<div class="dl-target-box">
                    <div class="dl-target-title">📄 Download Target — Single DCM</div>
                    <div class="dl-target-desc">
                        <b>Mode:</b> Single file only<br>
                        <b>File:</b> {b_sel}<br>
                        <b>Changes:</b> {n_staged} tag(s) will be modified<br>
                        <b>Output:</b> {b_base}_modified.dcm
                    </div></div>""", unsafe_allow_html=True)

            ac1, ac2 = st.columns([1, 2])
            with ac1:
                if st.button("🗑️ Clear All Staged", use_container_width=True):
                    st.session_state.cmp_staged         = {}
                    st.session_state.cmp_manual_staged  = {}
                    st.session_state.cmp_result_bytes   = None
                    st.session_state.cmp_result_zip     = None
                    st.session_state.cmp_result_summary = None
                    st.session_state.cmp_result_log     = None
                    st.rerun()
            with ac2:
                n_zip_files = len(st.session_state.cmp_b_zip_list) if is_zip_mode else 0
                btn_label   = (f"🚀 Apply to All {n_zip_files} Files & Create ZIP"
                               if is_zip_mode else "🚀 Apply & Download Single DCM")
                if st.button(btn_label, type="primary", use_container_width=True):
                    if is_zip_mode:
                        with st.spinner(f"Applying {n_staged} change(s) to all files..."):
                            rb, results, summary = apply_staged_to_zip(
                                st.session_state.cmp_b_zip_bytes, merged_staged)
                        st.session_state.cmp_result_zip     = rb
                        st.session_state.cmp_result_bytes   = None
                        st.session_state.cmp_result_summary = summary
                        st.session_state.cmp_result_log     = results
                    else:
                        with st.spinner("Applying changes to selected DCM..."):
                            rb, results = apply_modifications_to_ds(
                                st.session_state.cmp_ds_b, merged_staged)
                        st.session_state.cmp_result_bytes   = rb
                        st.session_state.cmp_result_zip     = None
                        st.session_state.cmp_result_summary = None
                        st.session_state.cmp_result_log     = results

            if st.session_state.cmp_result_bytes:
                b_base = Path(st.session_state.cmp_b_name).stem
                st.markdown('<div class="success-banner">✅ Modified File B ready for download!</div>',
                            unsafe_allow_html=True)
                st.download_button(
                    label="⬇️ Download Modified File B (.dcm)",
                    data=st.session_state.cmp_result_bytes,
                    file_name=f"{b_base}_modified.dcm",
                    mime="application/octet-stream",
                    use_container_width=True, type="primary")

            if st.session_state.cmp_result_zip:
                s = st.session_state.cmp_result_summary
                st.markdown(
                    f'<div class="success-banner">✅ Modified ZIP ready! &nbsp;'
                    f"Processed: <b>{s['success']}</b> &nbsp;·&nbsp; "
                    f"Skipped: <b>{s['skipped']}</b> &nbsp;·&nbsp; "
                    f"Errors: <b>{s['errors']}</b></div>",
                    unsafe_allow_html=True)
                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("📁 Total",     s["total"])
                mc2.metric("✅ Processed", s["success"])
                mc3.metric("⏭️ Skipped",   s["skipped"])
                mc4.metric("❌ Errors",    s["errors"])
                zip_out_name = st.session_state.cmp_b_name.replace(".zip", "_modified.zip")
                st.download_button(
                    label=f"⬇️ Download Modified ZIP ({s['success']} files)",
                    data=st.session_state.cmp_result_zip,
                    file_name=zip_out_name, mime="application/zip",
                    use_container_width=True, type="primary")
                if st.session_state.cmp_result_log:
                    with st.expander("📊 Modification Report", expanded=False):
                        st.dataframe(pd.DataFrame(st.session_state.cmp_result_log),
                                     use_container_width=True, height=300, hide_index=True)

            if st.session_state.cmp_result_bytes or st.session_state.cmp_result_zip:
                log_csv = pd.DataFrame(staged_rows).to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Download Change Log CSV",
                    data=log_csv, file_name="change_log.csv",
                    mime="text/csv", use_container_width=True)

        st.markdown("---")
        st.markdown("### 📥 Export Full Diff Report")
        ex1, ex2 = st.columns(2)
        with ex1:
            st.download_button(
                label="⬇️ Download Full Diff CSV",
                data=df_full.to_csv(index=False).encode("utf-8"),
                file_name="dicom_diff.csv", mime="text/csv",
                use_container_width=True)
        with ex2:
            if OPENPYXL_OK:
                st.download_button(
                    label="⬇️ Download Full Diff Excel",
                    data=make_excel_diff(df_full),
                    file_name="dicom_diff.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            else:
                st.info("💡 `pip install openpyxl` to enable Excel export.")

# ══════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:16px 0 20px;">
        <div style="width:56px;height:56px;margin:0 auto 10px;
            display:flex;align-items:center;justify-content:center;">
            {sidebar_logo_html}
        </div>
        <div style="font-size:14px;font-weight:800;letter-spacing:2px;
            background:linear-gradient(90deg,#00d4ff,#0066ff);
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;">
            SwiftMR</div>
        <div style="font-size:11px;color:#8892a4;margin-top:2px;letter-spacing:1px;">
            DICOM Header Editor</div>
    </div>""", unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">📖 How to Use</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="mode-label">Validation Mode</div>
    <div class="how-step"><div class="how-step-num">1</div>
        <div class="how-step-text">Upload <code>.dcm</code> or <code>.zip</code></div></div>
    <div class="how-step"><div class="how-step-num">2</div>
        <div class="how-step-text">Search &amp; browse tags (paginated)</div></div>
    <div class="how-step"><div class="how-step-num">3</div>
        <div class="how-step-text">Select tag → Enter value → Queue</div></div>
    <div class="how-step"><div class="how-step-num">4</div>
        <div class="how-step-text">Check Download Target → Apply &amp; Download</div></div>
    <div class="mode-label" style="margin-top:14px;">Compare Mode</div>
    <div class="how-step"><div class="how-step-num">1</div>
        <div class="how-step-text">Upload File A + File B (.dcm or .zip)</div></div>
    <div class="how-step"><div class="how-step-num">2</div>
        <div class="how-step-text">Run Comparison → View Diff Table</div></div>
    <div class="how-step"><div class="how-step-num">3</div>
        <div class="how-step-text">Inline Edit (diff) or Manual Edit (any tag)</div></div>
    <div class="how-step"><div class="how-step-num">4</div>
        <div class="how-step-text">Apply → Single DCM or Full ZIP download</div></div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">⚠️ Notes</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="note-item"><span class="note-icon">🔒</span>
        <span>Original files are <b>never modified</b></span></div>
    <div class="note-item"><span class="note-icon">🛡️</span>
        <span>Pixel Data <b>(7FE0,0010)</b> is protected</span></div>
    <div class="note-item"><span class="note-icon">💾</span>
        <span>Transfer Syntax &amp; Meta Header <b>fully preserved</b></span></div>
    <div class="note-item"><span class="note-icon">📦</span>
        <span>ZIP saved <b>without compression</b> (DICOM safe)</span></div>
    <div class="note-item"><span class="note-icon">🚫</span>
        <span>Do <b>not</b> upload real patient data (PHI)</span></div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">🔧 Common Tags</div>', unsafe_allow_html=True)
    tab1, tab2, tab3 = st.tabs(["Standard", "AIRS Upload", "AIRS Recon"])

    with tab1:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(0020,000E)</span>
            <span class="tag-name">SeriesInstanceUID</span></div>
        <div class="tag-row"><span class="tag-code">(0020,000D)</span>
            <span class="tag-name">StudyInstanceUID</span></div>
        <div class="tag-row"><span class="tag-code">(0008,103E)</span>
            <span class="tag-name">SeriesDescription</span></div>
        <div class="tag-row"><span class="tag-code">(0010,0010)</span>
            <span class="tag-name">PatientName</span></div>
        <div class="tag-row"><span class="tag-code">(0008,0060)</span>
            <span class="tag-name">Modality</span></div>
        <div class="tag-row"><span class="tag-code">(2001,9000)</span>
            <span class="tag-name">Private (Philips)</span></div>
        """, unsafe_allow_html=True)

    with tab2:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(00E1,1010)</span>
            <span class="tag-name">StudyId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1011)</span>
            <span class="tag-name">SeriesId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1012)</span>
            <span class="tag-name">ImageId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1014)</span>
            <span class="tag-name">UploadId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1015)</span>
            <span class="tag-name">DeviceId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1030)</span>
            <span class="tag-name">DispatchUnitId</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1031)</span>
            <span class="tag-name">PostProcessingType</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1032)</span>
            <span class="tag-name">SourceDispatchUnitId</span><span class="tag-vr">SH</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1033)</span>
            <span class="tag-name">ADCType</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1034)</span>
            <span class="tag-name">ADCNoiseThreshold</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1035)</span>
            <span class="tag-name">BValue</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1036)</span>
            <span class="tag-name">IsMarked</span><span class="tag-vr">CS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1037)</span>
            <span class="tag-name">NumberOfProjections</span><span class="tag-vr">IS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1038)</span>
            <span class="tag-name">RadialAngle</span><span class="tag-vr">IS</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1039)</span>
            <span class="tag-name">Zoom</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1040)</span>
            <span class="tag-name">SliceOrder</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1041)</span>
            <span class="tag-name">Orientation</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1042)</span>
            <span class="tag-name">Gap</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1043)</span>
            <span class="tag-name">Thickness</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1044)</span>
            <span class="tag-name">CalculatedBvalue</span><span class="tag-vr">SL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1045)</span>
            <span class="tag-name">PostprocessingMode</span><span class="tag-vr">LO</span></div>
        """, unsafe_allow_html=True)

    with tab3:
        st.markdown("""
        <div class="tag-row"><span class="tag-code">(00E1,1020)</span>
            <span class="tag-name">InputSnr</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1021)</span>
            <span class="tag-name">OutputSnr</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1022)</span>
            <span class="tag-name">DenoisingLevel</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1023)</span>
            <span class="tag-name">Sharpness</span><span class="tag-vr">FL</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1024)</span>
            <span class="tag-name">ModelPath</span><span class="tag-vr">OB</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1025)</span>
            <span class="tag-name">ModelHeader</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1026)</span>
            <span class="tag-name">ModelVersion</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1034)</span>
            <span class="tag-name">ADCNoiseThreshold</span><span class="tag-vr">LO</span></div>
        <div class="tag-row"><span class="tag-code">(00E1,1035)</span>
            <span class="tag-name">BValue</span><span class="tag-vr">LO</span></div>
        """, unsafe_allow_html=True)

    st.divider()
    st.markdown("""
    <div style="text-align:center;font-size:11px;color:#4a5568;padding:8px 0;">
        © 2026 AIRS Medical Inc.<br>All rights reserved.<br>Global Technical Support
    </div>""", unsafe_allow_html=True)
